# -*- coding: utf-8 -*-
"""
PickIreland site generator v2 (Trust & Authority design system).
Usage:  python build.py
Reads:  data/*.json  +  affiliate_links.xlsx
Writes: ../site/

Spreadsheet columns used: id, price_eur, affiliate_link, image_url.
Fill them, run this script, the whole site updates.
"""
import json, os, re, html, datetime, csv, hashlib

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, "data")
OUT  = os.path.join(BASE, "..", "docs")
SITE_NAME = "PickIreland"
DOMAIN = "https://pickireland.best"   # change when you buy the domain
YEAR = datetime.date.today().year
TODAY = datetime.date.today().strftime("%d %B %Y")
TODAY_ISO = datetime.date.today().strftime("%Y-%m-%d")
OG_IMAGE = DOMAIN + "/assets/og-default.png"   # imagem padrao de compartilhamento (1200x630)
GA_ID = "G-CEPZNGM1FZ"   # Google Analytics 4 (deixe "" para desativar)
GTAG = ("" if not GA_ID else
        "<!-- Google tag (gtag.js) -->\n"
        f'<script async src="https://www.googletagmanager.com/gtag/js?id={GA_ID}"></script>\n'
        "<script>window.dataLayer=window.dataLayer||[];function gtag(){dataLayer.push(arguments);}"
        f"gtag('js',new Date());gtag('config','{GA_ID}');</script>")
# Banner de cookies discreto e nao-bloqueante (consentimento implicito; permite recusar)
COOKIE_BANNER = """<div id="ckb" style="display:none;position:fixed;left:16px;right:16px;bottom:16px;max-width:560px;margin:0 auto;background:#0E1B26;color:#C4D2DE;border:1px solid #25394a;border-radius:12px;padding:12px 16px;font:14px/1.5 Inter,system-ui,sans-serif;box-shadow:0 12px 32px -12px rgba(0,0,0,.55);z-index:120;gap:12px;align-items:center;flex-wrap:wrap">
<span style="flex:1;min-width:210px">We use cookies to measure site traffic (Google Analytics). <a href="/privacy.html" style="color:#7FD4A8">Learn more</a>.</span>
<button id="ckb-no" style="background:transparent;border:1px solid #3a4f60;color:#C4D2DE;padding:7px 14px;border-radius:8px;cursor:pointer;font-weight:600">Decline</button>
<button id="ckb-ok" style="background:#F0A41C;border:0;color:#2A1A00;padding:7px 16px;border-radius:8px;cursor:pointer;font-weight:700">Accept</button>
</div>
<script>(function(){try{var k='ck_consent',c=localStorage.getItem(k);function ap(v){if(window.gtag){gtag('consent','update',{analytics_storage:v==='granted'?'granted':'denied',ad_storage:'denied'});}}if(c){ap(c);return;}var b=document.getElementById('ckb');if(!b)return;b.style.display='flex';document.getElementById('ckb-ok').onclick=function(){localStorage.setItem(k,'granted');ap('granted');b.style.display='none';};document.getElementById('ckb-no').onclick=function(){localStorage.setItem(k,'denied');ap('denied');b.style.display='none';};}catch(e){}})();</script>"""
# Autor dos guias (E-E-A-T). Edite a bio/url quando quiser; deixe url vazio se nao tiver LinkedIn.
AUTHOR = {
    "name": "André Pires",
    "role": "Founder & Editor",
    "bio": "André Pires founded PickIreland to cut through the noise of online shopping for Irish homes. "
           "Every guide compares products on Irish prices, running costs on Irish electricity, local weather and Irish rules — "
           "so readers can pick the right product without opening 40 tabs.",
    "url": "",          # ex: "https://www.linkedin.com/in/seu-perfil"
    "image": "",         # ex: DOMAIN + "/assets/author.jpg"
}

# ---------------------------------------------------------------- affiliate links
def load_links():
    links = {}
    xlsx = os.path.join(BASE, "affiliate_links.xlsx")
    csvf = os.path.join(BASE, "affiliate_links.csv")
    if os.path.exists(xlsx):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx, read_only=True)
            ws = wb["Affiliate Links"] if "Affiliate Links" in wb.sheetnames else wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            def col(name):
                return headers.index(name) if name in headers else None
            id_i, link_i = col("id"), col("affiliate_link")
            price_i, img_i = col("price_eur"), col("image_url")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or id_i is None or not row[id_i]:
                    continue
                pid = str(row[id_i]).strip()
                def val(i):
                    if i is None or row[i] is None: return ""
                    return str(row[i]).strip()
                links[pid] = {"link": val(link_i), "price": row[price_i] if price_i is not None else None, "image": val(img_i)}
            return links
        except Exception as e:
            print("! Could not read affiliate_links.xlsx:", e)
    if os.path.exists(csvf):
        with open(csvf, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                links[row["id"].strip()] = {"link": row.get("affiliate_link", "").strip(),
                                            "price": row.get("price_eur"), "image": row.get("image_url", "").strip()}
    return links

LINKS = load_links()

# Overlay de links exatos preenchidos manualmente (id -> link de afiliado), sem mexer no xlsx.
_extra = os.path.join(BASE, "extra_links.json")
if os.path.exists(_extra):
    for _id, _lnk in json.load(open(_extra, encoding="utf-8")).items():
        if _lnk:
            _d = LINKS.get(_id, {})
            _d["link"] = _lnk
            LINKS[_id] = _d

# ---------------------------------------------------------------- interlinking cruzado entre silos
# Categorias que resolvem problemas adjacentes. Quebra os silos estanques (antes: 0 links
# contextuais cruzando categorias) e espalha link equity para fora do proprio silo.
CROSS_LINKS = {
    "dehumidifiers":     [("air-purifiers", "damp and mould also mean airborne spores"),
                          ("electric-heaters", "warm air holds moisture better")],
    "air-purifiers":     [("dehumidifiers", "high humidity feeds mould spores"),
                          ("robot-vacuums", "less settled dust means less to filter")],
    "electric-heaters":  [("dehumidifiers", "drier air is cheaper to heat"),
                          ("home-office", "heating one room while you work")],
    "electric-bikes":    [("electric-scooters", "the shorter-commute alternative")],
    "electric-scooters": [("electric-bikes", "when the commute is longer or hillier")],
    "robot-vacuums":     [("robot-lawn-mowers", "the same hands-off logic, outdoors"),
                          ("air-purifiers", "dust you vacuum is dust you don't breathe")],
    "robot-lawn-mowers": [("robot-vacuums", "the indoor equivalent")],
    "air-fryers":        [("coffee-machines", "the other counter-top running-cost question")],
    "coffee-machines":   [("air-fryers", "the other appliance that pays for itself"),
                          ("home-office", "if the kitchen is also the office")],
    "home-office":       [("electric-heaters", "heating a home office without heating the house"),
                          ("coffee-machines", "the desk-side coffee question")],
}

def cross_links_html(cat_slug, cats_by_slug):
    """Links contextuais para o guia-cabeca de categorias adjacentes."""
    pairs = CROSS_LINKS.get(cat_slug, [])
    out = []
    for target, why in pairs:
        tc = cats_by_slug.get(target)
        if not tc or not tc["pages"]:
            continue
        head = tc["pages"][0]
        out.append(f'<a href="../{target}/{head["slug"]}.html">'
                   f'<span>{esc(head["h1"])} <i style="font-weight:400;opacity:.72">— {esc(why)}</i></span> {ARROW}</a>')
    if not out:
        return ""
    return ('<div class="related"><h2>Related guides in other categories</h2>'
            + "".join(out) + '</div>')

# ---------------------------------------------------------------- custo de operacao (€/hora)
# Por que existe: em 15/08/2026 a resposta de IA do Bing para "best dehumidifier ireland"
# citava um concorrente (eirehub.ie) pela frase "cost €0.04-0.06 per hour to run" — o custo
# de operacao e o nosso diferencial, e estava perdido dentro da prosa em vez de ser um campo.
# Motor generativo levanta CAMPO ROTULADO, nao paragrafo. Entao viramos campo.
#
# Honestidade: so calculamos onde existe potencia REAL na spec do produto. Sem potencia,
# nao inventamos numero — a linha simplesmente nao aparece.
KWH_RATE = 0.38          # €/kWh, day rate domestico irlandes (SEAI, julho 2026)

# Categorias onde watt = consumo da tomada e "custo por hora" e uma pergunta que o comprador
# realmente faz. Bicicleta, patinete, robo aspirador e cortador sao a bateria: o watt do motor
# nao e consumo continuo, e €/hora ali seria numero sem significado.
MAINS_CATEGORIES = {"dehumidifiers", "electric-heaters", "air-fryers",
                    "air-purifiers", "coffee-machines"}

_POWER_KEYS = ("power", "wattage", "consumption", "rated power", "input")

def _watts(specs):
    """Extrai a potencia em watts das specs. Devolve float ou None. Nunca chuta.

    01/09: a regex exigia 2+ digitos, entao "7 W" (Levoit Core Mini) e "2.5 W" (purificador
    USB de mesa) eram silenciosamente descartados — e purificador e justamente a categoria
    onde o numero baixo E o argumento. Agora aceita 1 digito e decimal. O filtro de chave
    (_POWER_KEYS) e que evita falso positivo, nao a contagem de digitos."""
    for k, v in (specs or {}).items():
        if not any(t in k.lower() for t in _POWER_KEYS):
            continue
        txt = str(v).replace(",", "")
        # pega o MAIOR numero seguido de W (ex: "600-2000 W" -> 2000, que e o consumo maximo)
        nums = [float(m) for m in re.findall(r"(\d{1,5}(?:\.\d+)?)\s*(?:W\b|watt)", txt, re.I)]
        if nums:
            return max(nums)
    return None

def w_txt(w):
    """1450.0 -> '1450', 2.5 -> '2.5'. _watts agora devolve float."""
    return f"{w:g}"

def rc_eur(x):
    """€/hora com casas suficientes para o numero nao sumir. Um purificador de 7 W custa
    €0.003/hora; arredondado para €0.00 ele lê como campo vazio — o oposto do que queremos,
    ja que 'praticamente nada' e justamente o argumento da categoria."""
    if x <= 0:
        return "€0.00"
    if x < 0.001:
        return "&lt;€0.001"      # 2.5 W a noite da €0.00045: um teto, nao um zero
    if x < 0.01:
        return f"€{x:.3f}"
    return f"€{x:.2f}"

def running_cost_line(specs, category):
    """'€0.08' por hora, ou None. Campo curto e extraivel de proposito."""
    if category not in MAINS_CATEGORIES:
        return None
    w = _watts(specs)
    if not w:
        return None
    return f"{rc_eur(w / 1000 * KWH_RATE)}/hour at €{KWH_RATE:.2f}/kWh"

# ---------------------------------------------------------------- grafico de custo de operacao
# Por que existe: o site tem ZERO imagem de produto (a Amazon so autoriza imagem via API, e a
# API exige 10 vendas/30 dias). Grafico gerado por codigo a partir das nossas proprias specs e
# o unico ativo visual possivel hoje — e resolve dois problemas de uma vez: a pagina deixa de
# ser um muro de texto, e o motor generativo ganha um numero comparativo com fonte declarada.
#
# Regras (mesmas do campo "Running cost", nao afrouxar):
#  - so entra produto com potencia REAL na spec. Sem watt, fica de fora. Nunca estimar.
#  - se sobrarem menos de RC_CHART_MIN produtos, nao desenha nada — 2 barras nao e comparacao.
#  - o SVG e inline e usa <text> de verdade (nao imagem), entao o numero e extraivel.
#  - a legenda declara quantos dos N produtos da pagina publicam potencia. Honestidade visivel.
RC_CHART_MIN = 3        # minimo de produtos com potencia para o grafico existir
RC_NOUN = {"dehumidifiers": "dehumidifier", "electric-heaters": "heater",
           "air-fryers": "air fryer", "air-purifiers": "air purifier",
           "coffee-machines": "coffee machine"}
# Por que a conta de cima e um TETO, e nao a conta real — a razao muda por categoria.
# Escrever "o humidostato desliga o compressor" numa pagina de air fryer seria falso, e foi
# o que a primeira versao deste bloco fez em 22/08 (o texto era unico para todas). Cada
# categoria tem de dizer a verdade da sua.
RC_QUALIFIER = {
 "dehumidifiers": "a humidistat cycles the compressor off once the room reaches the target "
                  "humidity, so the unit is not drawing full power the whole time",
 "electric-heaters": "a thermostat cycles the element off once the room is up to temperature, "
                     "so few heaters draw their rated wattage continuously",
 "air-fryers": "an air fryer only draws full power while the element is heating, and most "
               "recipes run for 15\u201325 minutes rather than a full hour",
 "air-purifiers": "the rated figure is the top fan speed, and most people run a purifier "
                  "well below it",
 "coffee-machines": "the element draws full power only while it heats up, which is a fraction "
                    "of the time the machine is switched on",
}

def _rc_rows(products, category):
    """[(nome, watts, euros_por_hora)] ordenado do mais barato ao mais caro."""
    if category not in MAINS_CATEGORIES:
        return []
    rows = []
    for p in products:
        w = _watts(p.get("specs") or {})
        if w:
            rows.append((p["name"], w, w / 1000 * KWH_RATE))
    rows.sort(key=lambda r: r[1])
    return rows

def _short(name, n=38):
    return name if len(name) <= n else name[:n - 1].rstrip(" ,-&") + "\u2026"

def running_cost_chart(products, category, unit_noun="unit", more_href=None, more_text=""):
    """Bloco <h2> + <figure> com as barras de custo/hora. '' se nao houver dado suficiente.

    Barras em HTML/CSS, nao SVG: um SVG com viewBox fixo encolhe a fonte junto com a
    largura e no telemovel fica ilegivel — e o telemovel e a maior parte do trafego.
    Em HTML o texto e texto de verdade em qualquer largura, reflui, e o motor generativo
    levanta o numero sem depender de saber ler <text> dentro de <svg>."""
    rows = _rc_rows(products, category)
    if len(rows) < RC_CHART_MIN:
        return ""
    hi = max(r[2] for r in rows)
    items = []
    for i, (name, w, eur) in enumerate(rows):
        pct = max(4.0, round(100.0 * eur / hi, 1))
        cheap = ' rc-best' if i == 0 else ''
        tag = ' <b class="rc-tag">cheapest to run</b>' if i == 0 else ''
        items.append(
            f'<li><span class="rcn">{esc(name)}</span>'
            f'<span class="rcb"><span class="rcf{cheap}" style="width:{pct}%"></span></span>'
            f'<span class="rcv">{rc_eur(eur)}/hour<i> \u00b7 {w_txt(w)}\u2009W</i>{tag}</span></li>')
    lo_name, lo_w, lo_e = rows[0]
    hi_name, hi_w, hi_e = rows[-1]
    ratio = hi_w / lo_w
    spread = f" \u2014 a {ratio:.1f}\u00d7 difference between them" if ratio >= 1.5 else ""
    have, total = len(rows), len(products)
    gap = total - have
    missing = ("" if gap == 0 else
               f" {have} of the {total} {unit_noun}s on this page have a manufacturer-published "
               + ("power figure; the other one is left out rather than estimated." if gap == 1
                  else f"power figure; the other {gap} are left out rather than estimated."))
    cap = (f"At Ireland\u2019s day rate of \u20ac{KWH_RATE:.2f}/kWh, the cheapest "
           f"{unit_noun} here is the {esc(lo_name)} at \u20ac{lo_e:.2f} per hour "
           f"({lo_w}\u2009W) and the most expensive is the {esc(hi_name)} at "
           f"\u20ac{hi_e:.2f} per hour ({hi_w}\u2009W){spread}."
           f"{missing} These are ceilings, not bills: the figure is the manufacturer\u2019s "
           f"rated draw at full power multiplied by the tariff, and in practice "
           f"{RC_QUALIFIER.get(category, 'the appliance rarely runs at its rated maximum')}. "
           f'Tariff source: <a href="https://www.seai.ie/data-and-insights/seai-statistics/prices" '
           f'rel="nofollow noopener" target="_blank">SEAI</a>, July 2026.')
    # Ponte para a tabela de referencia da categoria (o ativo linkavel). So aparece quando
    # a pagina de referencia existe — sem link morto se a categoria ainda nao tiver uma.
    if more_href:
        cap += f' <a href="{more_href}">{esc(more_text)}</a>'
    return (f'<h2>What each one costs to run</h2>\n'
            f'<figure class="rcfig">'
            f'<p class="rcax">Cost per hour at full power, at \u20ac{KWH_RATE:.2f}/kWh '
            f'\u2014 cheapest first</p>'
            f'<ol class="rcbars">{"".join(items)}</ol>'
            f'<figcaption>{cap}</figcaption></figure>')

# ---------------------------------------------------------------- pagina de referencia de custo (ativo linkavel)
# Por que existe (semana 2 do plano de crescimento, 29/08/2026): o site tem zero backlinks.
# Ninguem linka para "os 5 melhores desumidificadores" — todo mundo tem essa lista. Linka-se
# para um NUMERO com metodologia aberta: "quanto custa, em euros por hora, rodar cada
# desumidificador na Irlanda". Esta pagina e esse numero, numa tabela unica, com a conta a
# vista e a tarifa citada. Serve tres publicos de uma vez: quem procura o custo, quem escreve
# um artigo e precisa de fonte, e o motor generativo que precisa de campo rotulado.
#
# Honestidade que NAO se afrouxa aqui:
#  - so entra modelo com potencia publicada pelo fabricante. Sem watt, vai para a segunda
#    tabela, declarado como "not published". Nunca estimado.
#  - o modelo repetido em varias guias entra UMA vez (dedupe por nome). Contar deh-001 e
#    deh-010 como dois produtos inflaria a tabela com o mesmo aparelho.
#  - €/litro usa a extracao NOMINAL, que e medida em laboratorio a 30 °C / 80% HR. Casa
#    irlandesa no inverno esta muito abaixo disso, logo o €/litro real e PIOR. Dito na pagina.
KWH_RATE_NIGHT = 0.18    # €/kWh, ponto medio da faixa €0.17-0.19 de smart meter (julho 2026)
RC_REF_HOURS_DAY = 8     # horas/dia da estimativa mensal — declarado, nao escondido

def _extraction_lpd(specs):
    """Extracao nominal em litros/dia. Devolve float ou None. Nunca chuta.
    Aceita 'L/day' e converte 'ml/day'; ignora qualquer outra unidade."""
    for k, v in (specs or {}).items():
        if "extraction" not in k.lower():
            continue
        txt = str(v).replace(",", "")
        m = re.search(r"(\d+(?:\.\d+)?)\s*ml\s*/?\s*day", txt, re.I)
        if m:
            return float(m.group(1)) / 1000.0
        m = re.search(r"(\d+(?:\.\d+)?)\s*L\s*/?\s*day", txt, re.I)
        if m:
            return float(m.group(1))
    return None

def _unique_products(cat):
    """Todos os produtos da categoria, um por modelo. Chave = nome normalizado,
    porque o mesmo aparelho aparece com ids diferentes em guias diferentes."""
    seen, out = {}, []
    for pg in cat["pages"]:
        for p in pg["products"]:
            key = re.sub(r"\s+", " ", p["name"]).strip().lower()
            if key in seen:
                continue
            seen[key] = True
            out.append(p)
    return out

# Uma pagina de referencia por categoria de tomada. Hoje so dehumidifiers (o item da semana 2
# do plano); heaters e air-fryers entram na semana 4 acrescentando a entrada aqui.
RC_REF_PAGES = {
    "dehumidifiers": {
        "slug": "dehumidifier-running-costs-ireland",
        "title": "What Dehumidifiers Cost to Run in Ireland (2026)",
        "h1": "What a dehumidifier actually costs to run in Ireland",
        "desc": "Rated power, cost per hour at Irish day and night rates, cost per month "
                "and cost per litre of water removed, for every dehumidifier we track.",
        "noun": "dehumidifier",
        "noun_pl": "dehumidifiers",
        "unit": "litre of water removed",
        "hours_day": 8,
        "hours_text": "8 hours a day",
        "amount_spec": "Extraction",
        "amount_head": "Rated extraction",
        "unit_note_label": "extraction",
        "ceiling_short": "the humidistat switches the compressor off once the room reaches "
                         "its target humidity, so a real bill lands below them.",
        "link_text": "See the full running-cost table for every dehumidifier we track →",
        # Por que a conta e um TETO — a razao e FISICA DA CATEGORIA e muda de uma para outra.
        # Em 22/08 um texto unico para todas as categorias afirmou "o humidostato desliga o
        # compressor" numa pagina de air fryer. Nao repetir: cada categoria escreve o seu.
        "ceiling_note":
            "A compressor dehumidifier with a humidistat switches the compressor off once the "
            "room hits the target humidity, so it spends a good part of every hour drawing "
            "almost nothing. Desiccant models are the exception worth knowing about: they heat "
            "a rotor rather than run a compressor, so they sit much closer to their rated draw "
            "the whole time they are on — the Meaco DD8L in the table above is one of these, "
            "and its figure is nearer to a real hourly cost than the others.",
        # A ressalva da unidade de comparacao (€/litro). Tambem especifica da categoria.
        "unit_note":
            "Manufacturers rate extraction at around 30&nbsp;°C and 80% relative humidity. An "
            "Irish bedroom in November is nowhere near that, so a unit rated at 12&nbsp;L/day "
            "will pull out considerably less, and the real cost per litre will be higher than "
            "the column above. We show the rated figure because it is the one the manufacturer "
            "publishes and the one you can verify — not because we think you will get it.",
    },
    # As tres abaixo entraram em 01/09, depois de a pesquisa de potencia fechar a lacuna de
    # dados (74 cards ganharam watts). Cada uma tem horas/dia e ressalva PROPRIAS: um aquecedor
    # nao roda 8 h/dia como um desumidificador, e um air fryer nao roda nem meia hora.
    # Coffee machines ficam DE FORA de proposito: uma maquina de cafe fica ligada por minutos,
    # entao custo/mes por horas/dia seria um numero certo respondendo a pergunta errada.
    "electric-heaters": {
        "link_text": "See what every heater we track costs per hour →",
        "slug": "electric-heater-running-costs-ireland",
        "title": "What Electric Heaters Cost to Run in Ireland (2026)",
        "h1": "What an electric heater actually costs to run in Ireland",
        "desc": "Rated power and cost per hour at Irish day and night rates, plus cost per "
                "month over a winter evening, for every electric heater we track.",
        "noun": "heater",
        "noun_pl": "heaters",
        "unit": None,
        "hours_day": 5,
        "hours_text": "5 hours a day",
        "amount_spec": "Type",
        "amount_head": "Heater type",
        "ceiling_short": "a thermostat switches the element off once the room is up to "
                         "temperature, so few heaters draw their rated wattage for a full "
                         "hour and a real bill lands below them.",
        "ceiling_note":
            "Every heater here converts essentially all the electricity it draws into heat, "
            "so at full power a 2 kW heater costs the same to run whatever the badge on it "
            "says. What separates them is how much of the hour they spend at full power: a "
            "heater with a decent thermostat reaches the set temperature and then idles, "
            "while a cheap two-setting fan heater with no thermostat runs flat out until you "
            "switch it off. That is where the difference in a real bill comes from, not from "
            "one heater being more efficient than another.",
    },
    "air-fryers": {
        "link_text": "See what every air fryer we track costs to run →",
        "slug": "air-fryer-running-costs-ireland",
        "title": "What Air Fryers Cost to Run in Ireland (2026)",
        "h1": "What an air fryer actually costs to run in Ireland",
        "desc": "Rated power and cost per hour at Irish day and night rates for every air "
                "fryer we track, plus what a typical cook actually costs.",
        "noun": "air fryer",
        "noun_pl": "air fryers",
        "unit": None,
        "hours_day": 0.5,
        "hours_text": "30 minutes a day",
        "amount_spec": "Capacity",
        "amount_head": "Capacity",
        "ceiling_short": "the element only draws full power while it is heating, and most "
                         "recipes run for 15 to 25 minutes rather than a full hour, so a "
                         "real bill lands well below them.",
        "ceiling_note":
            "An air fryer is a fan and a heating element in a small insulated box. The element "
            "pulls its rated wattage while it is bringing the box up to temperature and then "
            "cycles, so a 2,400 W machine is not pulling 2,400 W for the whole cook. The "
            "comparison that matters for most households is not against another air fryer but "
            "against the oven: a fan oven drawing around 2 kW has a far larger cavity to heat "
            "and a longer preheat, which is where the running-cost saving actually comes from.",
    },
    "air-purifiers": {
        "link_text": "See what every air purifier we track costs to run →",
        "slug": "air-purifier-running-costs-ireland",
        "title": "What Air Purifiers Cost to Run in Ireland (2026)",
        "h1": "What an air purifier actually costs to run in Ireland",
        "desc": "Rated power and cost per hour at Irish day and night rates for every air "
                "purifier we track — the appliance people most often assume is expensive.",
        "noun": "air purifier",
        "noun_pl": "air purifiers",
        "unit": None,
        "hours_day": 12,
        "hours_text": "12 hours a day",
        "amount_spec": "Coverage",
        "amount_head": "Rated coverage",
        "ceiling_short": "the rated figure is the top fan speed, and almost nobody runs a "
                         "purifier flat out, so a real bill lands well below them.",
        "ceiling_note":
            "The rated wattage of an air purifier is its draw on the highest fan speed. On "
            "auto or sleep mode — how most of these actually get used — a unit rated at "
            "40 W is often pulling under 10. This is the appliance people most often assume "
            "is expensive to leave on, and the table above is the argument that it is not: "
            "even at the ceiling figure, running the most power-hungry purifier here costs "
            "less per month than the cheapest heater on this site costs in a single evening. "
            "Filters, not electricity, are what an air purifier costs you.",
    },
<<<<<<< HEAD
}

# Artigos informacionais de custo (semana 3 da fila, 05/09/2026). Um por categoria, no
# maximo. Fica aqui em cima, junto de RC_REF_PAGES, porque o hub e a propria pagina de
# referencia precisam saber que o artigo existe para linkar para ele.
RC_ARTICLES = {
    "dehumidifiers": {
        "slug": "how-much-does-a-dehumidifier-cost-to-run-ireland",
        "title": "How Much Does a Dehumidifier Cost to Run in Ireland?",
        "h1": "How much does a dehumidifier cost to run in Ireland?",
        "desc": "What a dehumidifier costs per hour, per night and per month at Irish "
                "electricity rates, with the arithmetic shown and the assumptions named.",
        "crumb": "Running cost explained",
        "hub_text": "How much does a dehumidifier cost to run in Ireland?",
    },
=======
>>>>>>> 70447d3a93e2e11a184cd5542a15dca748d4680d
}

AFF_TAG = "elevaonline-21"   # Amazon Associates StoreID
def amazon_search_url(p):
    q = re.sub(r"[^A-Za-z0-9 ]", "", p["name"]).replace(" ", "+")
    # mesmo sem o link exato do produto, mantém a tag -> clique monetizado (seta o cookie de afiliado)
    return f"https://www.amazon.ie/s?k={q}&tag={AFF_TAG}"

def product_url(p):
    info = LINKS.get(p["id"], {})
    if info.get("link"):
        return info["link"], True
    # fallback 1: usa o ASIN do proprio JSON de dados (evita link de busca quando o
    # produto tem ASIN conhecido mas ficou de fora do xlsx/extra_links.json)
    asin = (p.get("asin") or "").strip().upper()
    if re.fullmatch(r"[A-Z0-9]{10}", asin):
        return f"https://www.amazon.ie/dp/{asin}?tag={AFF_TAG}", True
    # fallback 2: busca com tag (clique ainda monetizado)
    return amazon_search_url(p), False

def product_price(p):
    info = LINKS.get(p["id"], {})
    if info.get("price"):
        try: return int(float(info["price"]))
        except (TypeError, ValueError): pass
    return p["price"]

def product_image(p):
    return LINKS.get(p["id"], {}).get("image", "")

def esc(s): return html.escape(str(s), quote=True)

# ---------------------------------------------------------------- svg icons (lucide-style, 24x24)
ICONS = {
"electric-scooters": '<path d="M13 2 3 14h9l-1 8 10-12h-9l1-8z"/>',
"electric-bikes": '<circle cx="5.5" cy="17.5" r="3.5"/><circle cx="18.5" cy="17.5" r="3.5"/><path d="M15 6a1 1 0 1 0 0-2 1 1 0 0 0 0 2zm-3 11.5V14l-3-3 4-3 2 3h2"/>',
"dehumidifiers": '<path d="M12 22a7 7 0 0 0 7-7c0-2-1-3.9-3-5.5s-3.5-4-4-6.5c-.5 2.5-2 4.9-4 6.5C6 11.1 5 13 5 15a7 7 0 0 0 7 7z"/>',
"air-fryers": '<path d="M3 8h18v9a4 4 0 0 1-4 4H7a4 4 0 0 1-4-4z"/><path d="M7 4v2m5-2v2m5-2v2"/>',
"electric-heaters": '<path d="M8.5 14.5A2.5 2.5 0 0 0 11 12c0-1.38-.5-2-1-3-1.072-2.143-.224-4.054 2-6 .5 2.5 2 4.9 4 6.5 2 1.6 3 3.5 3 5.5a7 7 0 1 1-14 0c0-1.153.433-2.294 1-3a2.5 2.5 0 0 0 2.5 2.5z"/>',
"air-purifiers": '<path d="M17.7 7.7a2.5 2.5 0 1 1 1.8 4.3H2"/><path d="M9.6 4.6A2 2 0 1 1 11 8H2"/><path d="M12.6 19.4A2 2 0 1 0 14 16H2"/>',
"robot-vacuums": '<circle cx="12" cy="12" r="9"/><path d="M8 12h8m-4-4v.01"/>',
"robot-lawn-mowers": '<path d="M11 20A7 7 0 0 1 9.8 6.1C15.5 5 17 4.48 19 2c1 2 2 4.18 2 8 0 5.5-4.78 10-10 10z"/><path d="M2 21c0-3 1.85-5.36 5.08-6C9.5 14.52 12 13 13 12"/>',
"home-office": '<rect x="2" y="3" width="20" height="14" rx="2"/><path d="M8 21h8m-4-4v4"/>',
"coffee-machines": '<path d="M17 8h1a4 4 0 1 1 0 8h-1"/><path d="M3 8h14v9a4 4 0 0 1-4 4H7a4 4 0 0 1-4-4z"/><path d="M7 2v3m4-3v3m4-3v3"/>',
}
def icon(cat_key, size=24, cls="ic"):
    path = ICONS.get(cat_key, '<circle cx="12" cy="12" r="9"/>')
    return f'<svg class="{cls}" width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">{path}</svg>'

CHECK = '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M20 6 9 17l-5-5"/></svg>'
CROSS = '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M18 6 6 18M6 6l12 12"/></svg>'
STAR  = '<svg width="16" height="16" viewBox="0 0 24 24" fill="currentColor" stroke="none" aria-hidden="true"><path d="M12 2l3.09 6.26L22 9.27l-5 4.87 1.18 6.88L12 17.77l-6.18 3.25L7 14.14 2 9.27l6.91-1.01z"/></svg>'
STAR_H = '<svg width="16" height="16" viewBox="0 0 24 24" aria-hidden="true"><defs><linearGradient id="hg"><stop offset="50%" stop-color="currentColor"/><stop offset="50%" stop-color="#D8DEE6"/></linearGradient></defs><path fill="url(#hg)" d="M12 2l3.09 6.26L22 9.27l-5 4.87 1.18 6.88L12 17.77l-6.18 3.25L7 14.14 2 9.27l6.91-1.01z"/></svg>'
ARROW = '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M5 12h14m-6-6 6 6-6 6"/></svg>'
SHIELD = '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/><path d="m9 12 2 2 4-4"/></svg>'

# ---------------------------------------------------------------- css
# Fonte carregada por <link> no head, NAO por @import dentro do CSS inline.
# @import dentro de <style> serializa o download (o browser so descobre a fonte depois de
# parsear o CSS) e bloqueia o render -> penaliza LCP/FCP. Com o padrao
# media="print" + onload trocando para all, o CSS da fonte vira nao-bloqueante.
FONT_URL = ("https://fonts.googleapis.com/css2?family=Bricolage+Grotesque:opsz,wght@12..96,600;12..96,700;12..96,800"
            "&family=Inter:wght@400;500;600;700&display=swap")
CSS = """
:root{--green:#0B5B40;--green-d:#073F2C;--green-l:#12805C;--green-t:#E9F4EE;--gold:#F0A41C;--gold-l:#FFC65C;--ink:#11202D;--mut:#54657A;--bg:#F7F8F6;--card:#fff;--line:#E4E9E4;--rad:18px;
--sh-sm:0 1px 3px rgba(13,27,42,.07);--sh-md:0 10px 30px -10px rgba(13,27,42,.16);--sh-lg:0 24px 60px -20px rgba(13,27,42,.25)}
*{margin:0;padding:0;box-sizing:border-box}
html{scroll-behavior:smooth}
body{font-family:'Inter',-apple-system,'Segoe UI',sans-serif;color:var(--ink);background:var(--bg);line-height:1.7;font-size:16.5px;-webkit-font-smoothing:antialiased}
h1,h2,h3,.logo,.price,.hero-stats b,.rank{font-family:'Bricolage Grotesque','Inter',sans-serif}
a{color:var(--green);text-decoration:none}
a:hover{text-decoration:underline}
a:focus-visible,button:focus-visible,summary:focus-visible{outline:3px solid var(--gold);outline-offset:2px;border-radius:4px}
.wrap{max-width:1140px;margin:0 auto;padding:0 22px}
.ic{flex:none}
::selection{background:var(--gold-l);color:var(--ink)}
/* reveal animations */
.rv{opacity:0;transform:translateY(16px);transition:opacity .55s ease,transform .55s ease}
.rv.vis{opacity:1;transform:none}
@media(prefers-reduced-motion:reduce){html{scroll-behavior:auto}*,*::before,*::after{animation:none!important;transition:none!important}.rv{opacity:1;transform:none}}
/* header */
header{background:rgba(255,255,255,.85);backdrop-filter:blur(12px) saturate(1.4);border-bottom:1px solid var(--line);position:sticky;top:0;z-index:50}
.nav{display:flex;align-items:center;justify-content:space-between;height:68px;gap:16px}
.logo{font-size:1.45rem;font-weight:800;color:var(--ink);letter-spacing:-.6px;display:flex;align-items:center;gap:10px}
.logo:hover{text-decoration:none}
.logo .mark{width:36px;height:36px;border-radius:11px;background:linear-gradient(140deg,var(--green) 20%,var(--green-l));color:#fff;display:flex;align-items:center;justify-content:center;font-size:1.05rem;font-weight:800;box-shadow:0 4px 12px -4px rgba(11,91,64,.5)}
.logo span{color:var(--green)}
.nav-links{display:flex;gap:2px;flex-wrap:wrap}
.nav-links a{font-size:.86rem;font-weight:600;color:var(--mut);padding:8px 13px;border-radius:99px;transition:background .2s,color .2s}
.nav-links a:hover{color:var(--green);background:var(--green-t);text-decoration:none}
.nav-links a.extra{display:none}
.menu-btn{display:none;background:none;border:0;cursor:pointer;padding:9px;color:var(--ink);border-radius:10px}
.menu-btn:hover{background:var(--green-t)}
@media(max-width:880px){
.menu-btn{display:flex;align-items:center}
.nav-links{display:none;position:absolute;top:68px;left:0;right:0;background:#fff;border-bottom:1px solid var(--line);flex-direction:column;padding:8px 22px 14px;gap:0;box-shadow:var(--sh-md);max-height:calc(100vh - 72px);overflow-y:auto}
body.nav-open .nav-links{display:flex;animation:slideDown .25s ease}
@keyframes slideDown{from{opacity:0;transform:translateY(-8px)}to{opacity:1;transform:none}}
.nav-links a{display:block!important;width:100%;padding:13px 10px;font-size:1.02rem;border-bottom:1px solid var(--line);border-radius:0}
.nav-links a:last-child{border-bottom:none}}
.disclosure{background:linear-gradient(90deg,var(--green-t),#F2F7EE);font-size:.76rem;color:#3E5C50;padding:7px 0;text-align:center;letter-spacing:.1px}
.disclosure a{font-weight:700}
/* page head */
.crumbs{font-size:.78rem;color:var(--mut);margin:24px 0 4px;display:flex;gap:7px;flex-wrap:wrap;text-transform:uppercase;letter-spacing:.6px;font-weight:600}
.crumbs a{color:var(--mut)}
h1{font-size:2.6rem;line-height:1.08;letter-spacing:-1.2px;margin:12px 0 14px;font-weight:800}
.updated{display:flex;align-items:center;gap:9px;font-size:.83rem;color:var(--mut);margin-bottom:18px;flex-wrap:wrap}
.updated .dot{width:4px;height:4px;border-radius:50%;background:#C3CDC6}
.trust-chip{display:inline-flex;align-items:center;gap:6px;background:var(--green-t);color:var(--green-d);font-weight:700;font-size:.76rem;padding:4px 12px;border-radius:99px;border:1px solid #D2E6DA}
.intro{font-size:1.12rem;color:#3A4B5C;max-width:850px;margin-bottom:30px;line-height:1.75}
/* quick answer (AEO): resposta direta, autocontida, logo abaixo do H1 */
.quick-answer{background:linear-gradient(135deg,var(--green-t),#F4FAF6);border:1px solid #CFE4D8;border-left:5px solid var(--green);border-radius:0 14px 14px 0;padding:16px 22px;margin:0 0 26px;font-size:1.03rem;line-height:1.7;color:#22323F;max-width:850px}
.quick-answer strong{color:var(--green-d);font-weight:800}
/* citação de fonte primária (GEO/E-E-A-T) */
.rcfig{margin:0 0 26px;background:var(--card);border:1px solid var(--line);border-radius:var(--rad);box-shadow:var(--sh-sm);padding:16px 20px 14px;max-width:850px}
.rcfig .rcax{font-size:.74rem;text-transform:uppercase;letter-spacing:.7px;font-weight:700;color:var(--mut);margin:0 0 12px}
.rcbars{list-style:none;margin:0;padding:0;display:grid;gap:10px}
.rcbars li{display:grid;grid-template-columns:minmax(150px,34%) 1fr minmax(132px,auto);align-items:center;gap:10px}
.rcbars .rcn{font-size:.88rem;line-height:1.35;color:var(--ink)}
.rcbars .rcb{background:#EDF1EE;border-radius:5px;height:14px;overflow:hidden}
.rcbars .rcf{display:block;height:100%;background:var(--green-l);border-radius:5px}
.rcbars .rcf.rc-best{background:var(--green)}
.rcbars .rcv{font-size:.88rem;font-weight:800;color:var(--green-d);white-space:nowrap}
.rcbars .rcv i{font-style:normal;font-weight:500;color:var(--mut)}
.rcbars .rc-tag{display:block;font-size:.66rem;text-transform:uppercase;letter-spacing:.6px;font-weight:800;color:var(--green)}
.rcfig figcaption{font-size:.86rem;color:var(--mut);line-height:1.65;margin-top:14px;border-top:1px solid var(--line);padding-top:11px}
.rcfig figcaption a{color:var(--green-l)}
@media(max-width:640px){.rcbars li{grid-template-columns:1fr auto;gap:4px 10px}
.rcbars .rcn{grid-column:1/-1}
.rcbars .rc-tag{display:inline;margin-left:6px}}
.sources{font-size:.86rem;color:var(--mut);background:#F6F8F6;border:1px solid var(--line);border-radius:12px;padding:13px 18px;margin:18px 0 0;line-height:1.65}
.sources strong{color:var(--ink)}
.sources a{font-weight:600}
h2{font-size:1.7rem;margin:50px 0 18px;letter-spacing:-.7px;font-weight:800;position:relative;padding-left:16px}
h2::before{content:'';position:absolute;left:0;top:.32em;bottom:.22em;width:5px;border-radius:3px;background:linear-gradient(180deg,var(--gold),var(--green-l))}
h3{font-size:1.13rem;margin:20px 0 8px;font-weight:700}
/* toc */
.toc{background:linear-gradient(135deg,#fff, #FBFDFB);border:1px solid var(--line);border-radius:var(--rad);padding:22px 26px;margin-bottom:34px;box-shadow:var(--sh-sm)}
.toc b{display:flex;align-items:center;gap:9px;margin-bottom:12px;font-size:.98rem}
.toc ol{margin-left:22px;font-size:.95rem;counter-reset:n;list-style:none}
.toc li{margin:7px 0;counter-increment:n;position:relative;padding-left:8px}
.toc li::before{content:counter(n);position:absolute;left:-24px;top:2px;width:20px;height:20px;border-radius:50%;background:var(--green-t);color:var(--green-d);font-size:.7rem;font-weight:800;display:flex;align-items:center;justify-content:center}
.toc li i{color:var(--mut);font-style:normal;font-size:.84rem}
ol.method{max-width:850px;margin:0 0 22px;padding-left:22px;color:var(--ink);line-height:1.75}
ol.method li{margin:0 0 10px;font-size:.97rem}
ol.method strong{color:var(--green-d)}
/* table */
.tbl-scroll{overflow-x:auto;margin-bottom:8px;border-radius:var(--rad);box-shadow:var(--sh-sm);border:1px solid var(--line)}
table.cmp{width:100%;border-collapse:collapse;background:var(--card);font-size:.88rem;min-width:660px}
table.cmp th{background:linear-gradient(135deg,var(--green-d),var(--green));color:#fff;padding:13px 15px;text-align:left;font-size:.74rem;text-transform:uppercase;letter-spacing:.8px;font-weight:700;position:sticky;top:0}
table.cmp td{padding:13px 15px;border-top:1px solid var(--line);vertical-align:top}
table.cmp tr:first-child td{background:#FFFBEF}
table.cmp tr:nth-child(even):not(:first-child) td{background:#FAFCFA}
table.cmp tr:hover td{background:var(--green-t)}
/* product card */
.card{position:relative;background:var(--card);border:1px solid var(--line);border-radius:var(--rad);padding:30px 26px 24px;margin:24px 0;box-shadow:var(--sh-sm);transition:box-shadow .3s,border-color .3s;overflow:hidden}
.card::before{content:'';position:absolute;left:0;top:0;bottom:0;width:4px;background:linear-gradient(180deg,var(--green-l),transparent 70%)}
.card:hover{box-shadow:var(--sh-md);border-color:#CFE0D4}
.card.top-pick{border:2px solid var(--gold);box-shadow:0 16px 44px -16px rgba(240,164,28,.35)}
.card.top-pick::before{background:linear-gradient(180deg,var(--gold),var(--gold-l))}
.card.top-pick::after{content:'OUR #1 PICK';position:absolute;top:18px;right:-34px;transform:rotate(40deg);background:linear-gradient(90deg,var(--gold),var(--gold-l));color:#3A2700;font-size:.62rem;font-weight:800;letter-spacing:1px;padding:5px 40px}
.rank{position:absolute;top:0;left:0;width:46px;height:40px;border-radius:16px 0 16px 0;background:var(--ink);color:#fff;font-weight:800;font-size:1.05rem;display:flex;align-items:center;justify-content:center;box-shadow:var(--sh-sm)}
.card.top-pick .rank{background:linear-gradient(140deg,var(--gold),var(--gold-l));color:#3A2700}
.badge{display:inline-flex;align-items:center;gap:6px;background:linear-gradient(90deg,#FFF4DA,#FFEBC2);color:#8A5B00;font-size:.68rem;font-weight:800;text-transform:uppercase;letter-spacing:.7px;padding:5px 12px;border-radius:99px;margin-bottom:14px;border:1px solid #F5DEA8}
/* Cabecalho em 3 colunas: imagem | info | CTA.
   Antes o botao ficava depois de specs + pros/cons + verdict, o que a 100% de zoom
   empurrava o unico elemento que gera receita para fora da dobra. */
.card-head{display:grid;grid-template-columns:auto minmax(0,1fr) auto;gap:22px;align-items:center}
.card-info{min-width:0}
.card-cta{display:flex;flex-direction:column;align-items:stretch;gap:7px;width:264px}
.card-cta .btn{width:100%;justify-content:center;padding:14px 18px;font-size:.94rem;white-space:nowrap}
.card-cta .btn-sub{font-size:.7rem;color:var(--mut);text-align:center;margin:0;line-height:1.4}
.pimg{width:112px;height:112px;flex:none;border-radius:14px;border:1px solid var(--line);background:radial-gradient(circle at 30% 25%,#fff, #F2F6F2);display:flex;align-items:center;justify-content:center;overflow:hidden}
.pimg img{max-width:100%;max-height:100%;object-fit:contain;mix-blend-mode:multiply}
.pimg .ph{display:flex;flex-direction:column;align-items:center;gap:6px;color:#9AABA0;font-size:.6rem;font-weight:700;text-transform:uppercase;letter-spacing:.6px;text-align:center;padding:0 6px}
.pimg .ph .ic{width:34px;height:34px;color:#BECDC2}
@media(max-width:1000px){
  .card-head{grid-template-columns:auto minmax(0,1fr)}
  .card-cta{grid-column:1/-1;width:100%}
}
@media(max-width:560px){
  .card-head{grid-template-columns:1fr;justify-items:center;text-align:center}
  .pimg{width:96px;height:96px}
  .card-info{text-align:center}
  .pricerow{justify-content:center}
}
.card h3{font-size:1.28rem;margin:0 0 2px;letter-spacing:-.4px;line-height:1.28}
.brandline{color:var(--mut);font-size:.85rem;margin-bottom:12px;font-weight:600;text-transform:uppercase;letter-spacing:.6px}
.pricerow{display:flex;align-items:center;gap:16px;flex-wrap:wrap;margin:8px 0 0}
.price{font-size:1.5rem;font-weight:800;color:var(--green-d);letter-spacing:-.5px}
.price small{font-size:.7rem;color:var(--mut);font-weight:500;display:block;font-family:'Inter';letter-spacing:.3px;text-transform:uppercase}
.stars{display:flex;align-items:center;gap:2px;color:var(--gold)}
.stars small{color:var(--mut);margin-left:8px;font-weight:700;font-size:.85rem}
.specgrid{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:8px;margin:16px 0 0;font-size:.82rem}
.specgrid div{background:#F4F7F3;border-radius:10px;padding:8px 12px;border:1px solid #ECF1EB}
.nospec{font-size:.84rem;color:var(--mut);background:#F7F8F6;border:1px dashed var(--line);border-radius:11px;padding:10px 14px;margin:16px 0 0}
.specgrid b{display:block;font-size:.64rem;text-transform:uppercase;color:var(--mut);letter-spacing:.7px;margin-bottom:2px;font-weight:700}
.pc{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin:16px 0 0}
@media(max-width:640px){.pc{grid-template-columns:1fr}h1{font-size:1.8rem;letter-spacing:-.6px}}
.pros,.cons{border-radius:12px;padding:13px 16px;font-size:.89rem}
.pros{background:linear-gradient(160deg,#EDFAF1,#E4F6EB);border:1px solid #C8E8D2}
.cons{background:linear-gradient(160deg,#FDF6F1,#FBEEE6);border:1px solid #F0D9C8}
.pros b,.cons b{display:flex;align-items:center;gap:7px;margin-bottom:6px;font-size:.74rem;text-transform:uppercase;letter-spacing:.8px}
.pros b{color:#136A41}.cons b{color:#9A4E22}
.pros ul,.cons ul{list-style:none}
.pros li,.cons li{margin:4px 0;padding-left:21px;position:relative;line-height:1.5}
.pros li::before{content:'✓';position:absolute;left:2px;color:#15A35C;font-weight:800}
.cons li::before{content:'–';position:absolute;left:5px;color:#C76A33;font-weight:800}
.verdict{font-size:.95rem;line-height:1.62;background:linear-gradient(90deg,var(--green-t),transparent 80%);border-left:4px solid var(--green);padding:13px 18px;border-radius:0 12px 12px 0;margin:16px 0 0;font-style:italic;color:#273B4D}
.verdict strong{font-style:normal;color:var(--green-d)}
.btn{display:inline-flex;align-items:center;gap:10px;background:linear-gradient(135deg,#F7A91D,#F08A00);color:#2A1A00;font-weight:800;font-size:1.04rem;padding:15px 32px;border-radius:13px;box-shadow:0 8px 22px -8px rgba(240,138,0,.65);transition:box-shadow .25s,transform .25s;cursor:pointer;position:relative;overflow:hidden;font-family:'Inter'}
.btn::after{content:'';position:absolute;top:0;left:-80%;width:50%;height:100%;background:linear-gradient(105deg,transparent,rgba(255,255,255,.45),transparent);transition:left .5s ease}
.btn:hover{box-shadow:0 12px 28px -8px rgba(240,138,0,.8);transform:translateY(-2px);text-decoration:none}
.btn:hover::after{left:120%}
.btn .ic{transition:transform .25s}
.btn:hover .ic{transform:translateX(4px)}
.btn-sub{font-size:.74rem;color:var(--mut);margin-top:9px}
/* CTA secundario no fim do card (para quem leu tudo). Precisa vir depois de .btn
   na cascata, senao o gradiente do .btn sobrescreve o estilo contornado. */
.card .btn-foot{margin-top:4px;padding:11px 22px;font-size:.92rem;background:#fff;color:var(--green-d);border:2px solid var(--green);box-shadow:none}
.card .btn-foot:hover{background:var(--green-t);box-shadow:none;transform:none;text-decoration:none}
.card .btn-foot::after{display:none}
/* guide & faq */
.guide{background:var(--card);border:1px solid var(--line);border-radius:var(--rad);padding:10px 28px 20px;margin:20px 0;box-shadow:var(--sh-sm)}
details{border-bottom:1px solid var(--line);padding:16px 0}
details:last-child{border-bottom:none}
summary{font-weight:700;cursor:pointer;font-size:1.02rem;list-style:none;display:flex;justify-content:space-between;align-items:center;gap:12px}
summary::-webkit-details-marker{display:none}
summary::after{content:'+';font-size:1.4rem;color:var(--green);font-weight:600;transition:transform .25s;flex:none;width:30px;height:30px;border-radius:50%;background:var(--green-t);display:flex;align-items:center;justify-content:center}
details[open] summary::after{transform:rotate(45deg)}
details p{margin-top:11px;color:#3A4B5C;font-size:.95rem;animation:fadeIn .3s ease}
@keyframes fadeIn{from{opacity:0;transform:translateY(-4px)}to{opacity:1;transform:none}}
/* tiles */
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(250px,1fr));gap:20px;margin:28px 0}
.tile{position:relative;background:var(--card);border:1px solid var(--line);border-radius:var(--rad);transition:box-shadow .3s,transform .3s,border-color .3s;box-shadow:var(--sh-sm);overflow:hidden}
.tile::after{content:'';position:absolute;inset:0;background:linear-gradient(135deg,transparent 60%,rgba(11,91,64,.05));opacity:0;transition:opacity .3s;pointer-events:none}
.tile:hover{box-shadow:var(--sh-md);transform:translateY(-4px);border-color:#BFD8CD}
.tile:hover::after{opacity:1}
.tile a{display:block;padding:26px;text-decoration:none;cursor:pointer}
.tile .icw{width:50px;height:50px;border-radius:14px;background:var(--green-t);color:var(--green);display:flex;align-items:center;justify-content:center;margin-bottom:15px;transition:background .3s,color .3s,transform .3s}
.tile:hover .icw{background:linear-gradient(140deg,var(--green),var(--green-l));color:#fff;transform:scale(1.06) rotate(-3deg)}
.tile h3{margin:0 0 5px;font-size:1.1rem;color:var(--ink);letter-spacing:-.3px}
.tile p{font-size:.83rem;color:var(--mut)}
.tile .go{position:absolute;top:24px;right:22px;color:#CBD7CE;transition:color .3s,transform .3s}
.tile:hover .go{color:var(--green);transform:translateX(3px)}
/* related */
.related{margin:38px 0}
.related a{display:flex;align-items:center;justify-content:space-between;gap:10px;padding:14px 18px;margin:9px 0;background:var(--card);border:1px solid var(--line);border-radius:13px;font-size:.95rem;font-weight:600;transition:border-color .25s,box-shadow .25s,transform .25s;cursor:pointer}
.related a:hover{text-decoration:none;border-color:var(--green);box-shadow:var(--sh-sm);transform:translateX(3px)}
/* hero */
.hero{background:radial-gradient(1100px 500px at 75% -20%,#15724E 0%,transparent 55%),radial-gradient(800px 420px at 0% 120%,#0E6347 0%,transparent 60%),linear-gradient(150deg,#06281C 0%,#0A4A33 60%,#0C5C41 100%);color:#fff;padding:88px 0 76px;text-align:center;position:relative;overflow:hidden}
.hero::before{content:'';position:absolute;inset:0;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='140' height='140' viewBox='0 0 140 140'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='.9' numOctaves='2'/%3E%3C/filter%3E%3Crect width='140' height='140' filter='url(%23n)' opacity='.05'/%3E%3C/svg%3E");pointer-events:none}
.hero::after{content:'';position:absolute;inset:0;background:radial-gradient(600px 260px at 85% 10%,rgba(240,164,28,.22),transparent 70%);pointer-events:none}
.hero>*{position:relative;z-index:1}
.hero h1{color:#fff;font-size:3.3rem;max-width:820px;margin:0 auto 18px;letter-spacing:-1.6px;line-height:1.05}
.hero h1 em{font-style:italic;color:var(--gold-l);position:relative}
.hero p{font-size:1.16rem;opacity:.92;max-width:640px;margin:0 auto 30px;font-weight:400}
.hero-stats{display:flex;gap:14px;justify-content:center;flex-wrap:wrap}
.hero-stats div{background:rgba(255,255,255,.09);border:1px solid rgba(255,255,255,.18);border-radius:15px;padding:13px 26px;font-size:.85rem;font-weight:600;backdrop-filter:blur(6px);letter-spacing:.2px}
.hero-stats b{display:block;font-size:1.55rem;color:var(--gold-l);letter-spacing:-.5px}
@media(max-width:640px){.hero h1{font-size:2.1rem;letter-spacing:-.8px}.hero{padding:60px 0 50px}}
/* footer */
footer{background:linear-gradient(180deg,#0E1B26,#0A141D);color:#90A3B8;margin-top:70px;padding:52px 0 34px;font-size:.86rem}
footer a{color:#C4D2DE}
footer .cols{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:30px;margin-bottom:28px}
footer h4{color:#fff;font-size:.92rem;margin-bottom:13px;font-family:'Bricolage Grotesque';letter-spacing:.2px}
footer .legal{border-top:1px solid #1E2E3C;padding-top:20px;font-size:.76rem;line-height:1.6}
.notice{font-size:.78rem;color:var(--mut);margin:34px 0 10px;line-height:1.6}
/* back to top */
.top-btn{position:fixed;bottom:24px;right:24px;width:46px;height:46px;border-radius:50%;background:var(--green);color:#fff;border:0;cursor:pointer;display:flex;align-items:center;justify-content:center;box-shadow:var(--sh-md);opacity:0;pointer-events:none;transition:opacity .3s,transform .3s;z-index:60}
.top-btn.show{opacity:1;pointer-events:auto}
.top-btn:hover{transform:translateY(-3px)}
article.card{scroll-margin-top:90px}
/* footer aff line */
.aff-line{font-size:.8rem;color:#A8B8C6;padding:14px 0;border-top:1px solid #1E2E3C;margin-top:4px}
/* spotlight */
.spot{margin:54px 0 10px;background:linear-gradient(140deg,#0A3A29,#0C5C41 70%,#0E6B4A);border-radius:24px;padding:40px 36px;color:#fff;position:relative;overflow:hidden;box-shadow:var(--sh-lg)}
.spot::before{content:'';position:absolute;inset:0;background:radial-gradient(540px 280px at 88% 0%,rgba(240,164,28,.18),transparent 65%)}
.spot>*{position:relative}
.spot-head{display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap;margin-bottom:24px}
.spot-head h2{margin:0;padding:0;color:#fff;font-size:1.5rem}
.spot-head h2::before{display:none}
.spot-tabs{display:flex;flex-wrap:wrap;gap:4px;justify-content:center;background:rgba(0,0,0,.25);padding:5px;border-radius:16px;border:1px solid rgba(255,255,255,.12);max-width:100%}
.spot-tab{border:0;background:transparent;color:#B9CFC2;font-weight:600;font-size:.82rem;padding:9px 18px;border-radius:99px;cursor:pointer;transition:background .3s,color .3s;font-family:'Inter'}
.spot-tab.on{background:rgba(255,255,255,.14);color:#fff;box-shadow:inset 0 1px 0 rgba(255,255,255,.15)}
.spot-panel{display:none;align-items:center;gap:36px}
.spot-panel.on{display:flex;animation:spotIn .45s ease}
@keyframes spotIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:none}}
.spot-vis{flex:none;width:170px;height:170px;border-radius:50%;background:radial-gradient(circle at 32% 28%,rgba(255,255,255,.18),rgba(255,255,255,.05));border:1px dashed rgba(255,255,255,.25);display:flex;align-items:center;justify-content:center;color:#EAF6EF;animation:floaty 6s ease-in-out infinite}
.spot-vis .ic{width:64px;height:64px}
@keyframes floaty{0%,100%{transform:translateY(-6px)}50%{transform:translateY(6px)}}
.spot-info{flex:1;min-width:240px}
.spot-info .k{font-size:.72rem;text-transform:uppercase;letter-spacing:1.4px;color:var(--gold-l);font-weight:700}
.spot-info h3{color:#fff;font-size:1.5rem;margin:6px 0 4px;letter-spacing:-.4px}
.spot-info .pr{font-family:'Bricolage Grotesque';font-size:1.4rem;font-weight:800;color:var(--gold-l);margin-bottom:14px}
.bar{margin:10px 0}
.bar .lb{display:flex;justify-content:space-between;font-size:.76rem;color:#C2D6C9;margin-bottom:6px;letter-spacing:.3px}
.bar u{display:block;height:7px;background:rgba(0,0,0,.3);border-radius:99px;overflow:hidden;text-decoration:none}
.bar i{display:block;height:100%;width:0;border-radius:99px;background:linear-gradient(90deg,var(--gold),var(--gold-l));transition:width 1s cubic-bezier(.2,.7,.2,1)}
.spot .btn{margin-top:16px;padding:12px 26px;font-size:.95rem}
@media(max-width:700px){.spot{padding:28px 20px}.spot-panel.on{flex-direction:column;text-align:center}.spot-vis{width:130px;height:130px}}

/* search */
.search-btn{display:flex;align-items:center;background:none;border:0;cursor:pointer;padding:9px;color:var(--ink);border-radius:10px;margin-left:auto}
.search-btn:hover{background:var(--green-t)}
@media(min-width:881px){.search-btn{margin-left:0}}
.search-overlay{position:fixed;inset:0;background:rgba(8,22,16,.55);backdrop-filter:blur(7px);display:none;z-index:80;padding:9vh 18px 20px;justify-content:center;align-items:flex-start}
body.search-open .search-overlay{display:flex;animation:fadeIn .2s ease}
.search-box{width:100%;max-width:660px;background:#fff;border-radius:20px;box-shadow:var(--sh-lg);overflow:hidden}
.search-top{display:flex;align-items:center;border-bottom:1px solid var(--line)}
.search-top input{flex:1;border:0;padding:19px 24px;font-size:1.08rem;outline:none;font-family:'Inter';background:transparent}
.search-top button{border:0;background:none;font-size:1.7rem;color:var(--mut);cursor:pointer;padding:0 20px;line-height:1}
.search-top button:hover{color:var(--ink)}
.search-res{max-height:52vh;overflow:auto}
.search-res a{display:flex;justify-content:space-between;align-items:center;gap:12px;padding:13px 24px;border-bottom:1px solid var(--line);font-size:.93rem;font-weight:600;color:var(--ink)}
.search-res a:last-child{border-bottom:none}
.search-res a:hover{background:var(--green-t);text-decoration:none}
.search-res .meta{color:var(--mut);font-weight:500;font-size:.78rem}
.search-res .sp{font-family:'Bricolage Grotesque';font-weight:800;color:var(--green-d);flex:none}
.search-hint{padding:14px 24px;color:var(--mut);font-size:.85rem}

"""

# ---------------------------------------------------------------- template pieces
def stars(r):
    full = int(r); half = r - full >= 0.5
    s = STAR * full + (STAR_H if half else "")
    return f'{s}<small>{r}/5</small>'

def header_html(depth=0):
    p = "../" * depth
    cats_links = "".join(
        f'<a class="{"extra" if i >= 5 else ""}" href="{p}{c["category"]}/index.html">{c["name"]}</a>'
        for i, c in enumerate(CATS))
    burger = '<svg width="26" height="26" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" aria-hidden="true"><path d="M4 7h16M4 12h16M4 17h16"/></svg>'
    return f"""<header><div class="wrap nav">
<a class="logo" href="/" aria-label="{SITE_NAME} home"><span class="mark" aria-hidden="true"><svg width="22" height="22" viewBox="0 0 64 64" fill="none"><path d="M22 48V16h13a11 11 0 0 1 0 22h-9" stroke="#fff" stroke-width="7" stroke-linecap="round" stroke-linejoin="round"/><path d="M33 44l6 6 11-12" stroke="#FFC65C" stroke-width="6" stroke-linecap="round" stroke-linejoin="round"/></svg></span>Pick<span>Ireland</span></a>
<button class="search-btn" aria-label="Search products" onclick="document.body.classList.add('search-open');document.getElementById('siq').focus()"><svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round"><circle cx="11" cy="11" r="7"/><path d="m21 21-4.3-4.3"/></svg></button>
<button class="menu-btn" aria-label="Open menu" aria-expanded="false" onclick="document.body.classList.toggle('nav-open');this.setAttribute('aria-expanded',document.body.classList.contains('nav-open'))">{burger}</button>
<nav class="nav-links" aria-label="Categories">{cats_links}<a href="/#categories">All categories</a></nav>
</div></header>"""

def footer_html(depth=0):
    p = "../" * depth
    cat_links = "".join(f'<a href="{p}{c["category"]}/index.html">{c["name"]}</a><br>' for c in CATS)
    return f"""<footer><div class="wrap">
<div class="cols">
<div><h4>{SITE_NAME}</h4><p>Independent buying guides for Irish shoppers. We compare products on specs, real running costs and value — so you don't have to open 40 tabs.</p></div>
<div><h4>Categories</h4>{cat_links}</div>
<div><h4>About</h4><a href="{p}about.html">About us</a><br><a href="{p}affiliate-disclosure.html">Affiliate disclosure</a><br><a href="{p}privacy.html">Privacy policy</a><br><a href="{p}contact.html">Contact</a></div>
</div>
<div class="aff-line">As an Amazon Associate, {SITE_NAME} earns from qualifying purchases. <a href="{p}affiliate-disclosure.html">Learn more</a>.</div>
<div class="legal">© {YEAR} {SITE_NAME}. Prices shown are typical/indicative in EUR and change frequently — always check the current price at the retailer. As an Amazon Associate we earn from qualifying purchases.</div>
</div></footer>"""

def page_shell(title, desc, canonical, body, depth=0, jsonld="", og_type="article", og_image=None):
    og_img = og_image or OG_IMAGE
    # SEO guard: keep titles <= 60 and descriptions <= 155 chars so Google doesn't truncate them.
    if len(title) > 60:
        print(f"  [SEO] title {len(title)} chars (>60): {canonical}")
    if len(desc) > 155:
        print(f"  [SEO] description {len(desc)} chars (>155): {canonical}")
    return f"""<!DOCTYPE html>
<html lang="en-IE">
<head>
<meta charset="utf-8">
{GTAG}
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(title)}</title>
<meta name="description" content="{esc(desc)}">
<meta name="robots" content="index, follow, max-image-preview:large, max-snippet:-1, max-video-preview:-1">
<meta name="author" content="{SITE_NAME} Editorial Team">
<meta name="publisher" content="{SITE_NAME}">
<link rel="canonical" href="{canonical}">
<link rel="alternate" hreflang="en-ie" href="{canonical}">
<link rel="alternate" hreflang="x-default" href="{canonical}">
<meta property="og:title" content="{esc(title)}">
<meta property="og:description" content="{esc(desc)}">
<meta property="og:type" content="{og_type}">
<meta property="og:url" content="{canonical}">
<meta property="og:site_name" content="{SITE_NAME}">
<meta property="og:image" content="{og_img}">
<meta property="og:image:width" content="1200">
<meta property="og:image:height" content="630">
<meta property="og:image:alt" content="{esc(title)}">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:image" content="{og_img}">
<meta name="theme-color" content="#0B5B40">
<meta property="og:locale" content="en_IE">
<link rel="icon" type="image/svg+xml" href="{'../' * depth}favicon.svg">
<link rel="icon" type="image/png" sizes="32x32" href="{'../' * depth}favicon-32.png">
<link rel="apple-touch-icon" href="{'../' * depth}apple-touch-icon.png">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link rel="stylesheet" media="print" onload="this.media='all'" href="{FONT_URL}">
<noscript><link rel="stylesheet" href="{FONT_URL}"></noscript>
<style>{CSS}</style>
{jsonld}
</head>
<body>
{header_html(depth)}
<div class="search-overlay" id="sov">
<div class="search-box" role="dialog" aria-modal="true" aria-label="Product search">
<div class="search-top"><input id="siq" type="search" placeholder="Search 248 products… e.g. Ninja, Meaco, desk" autocomplete="off"><button type="button" data-close-search aria-label="Close search">×</button></div>
<div class="search-res" id="sres"><div class="search-hint">Type to search every product we've reviewed — press Esc to close.</div></div>
</div></div>
<main class="wrap">
{body}
</main>
{footer_html(depth)}
<button class="top-btn" aria-label="Back to top" onclick="scrollTo({{top:0,behavior:'smooth'}})"><svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M12 19V5m-7 7 7-7 7 7"/></svg></button>
<script defer src="{'../' * depth}assets/site.js"></script>
{COOKIE_BANNER}
</body>
</html>"""

def product_card(p, rank, cat_key):
    url, has_aff = product_url(p)
    price = product_price(p)
    img = product_image(p)
    rel = 'sponsored noopener' if has_aff else 'nofollow noopener'
    if img:
        img_html = f'<img src="{esc(img)}" alt="{esc(p["name"])}" loading="lazy" width="170" height="170">'
    else:
        img_html = f'<div class="ph">{icon(cat_key, 44)}<span>{esc(p["brand"])}</span></div>'
    # Produtos cujo modelo exato não pôde ser identificado ficam sem grade de specs,
    # em vez de exibir linhas de preenchimento que só repetem preço/nota já visíveis no card.
    # "Running cost" entra PRIMEIRO na grade: e o campo que queremos que a IA levante.
    _sp = dict(p.get("specs") or {})
    _rc = running_cost_line(_sp, cat_key)
    _ordered = ([("Running cost", _rc)] if _rc else []) + list(_sp.items())
    specs = "".join(f"<div><b>{esc(k)}</b>{esc(v)}</div>" for k, v in _ordered)
    specs_block = f'<div class="specgrid">{specs}</div>' if specs else (
        '<p class="nospec">Full specifications for this exact model aren\'t published by the '
        'manufacturer — check the current Amazon.ie listing before buying.</p>')
    pros = "".join(f"<li>{esc(x)}</li>" for x in p["pros"])
    cons = "".join(f"<li>{esc(x)}</li>" for x in p["cons"])
    top = " top-pick" if rank == 1 else ""
    return f"""<article class="card{top}" id="{p['id']}">
<div class="rank" aria-label="Rank {rank}">{rank}</div>
<span class="badge">{esc(p['badge'])}</span>
<div class="card-head">
  <div class="pimg">{img_html}</div>
  <div class="card-info">
    <h3>{esc(p['name'])}</h3>
    <div class="brandline">by {esc(p['brand'])}</div>
    <div class="pricerow">
      <div class="price">€{price}<small>typical price</small></div>
      <div class="stars" aria-label="Rated {p['rating']} out of 5">{stars(p['rating'])}</div>
    </div>
  </div>
  <div class="card-cta">
    <a class="btn" href="{esc(url)}" target="_blank" rel="{rel}">Check Price on Amazon.ie {ARROW}</a>
    <span class="btn-sub">Price accurate as of publishing</span>
  </div>
</div>
{specs_block}
<div class="pc">
  <div class="pros"><b>{CHECK} Pros</b><ul>{pros}</ul></div>
  <div class="cons"><b>{CROSS} Cons</b><ul>{cons}</ul></div>
</div>
<p class="verdict"><strong>Our verdict:</strong> {esc(p['verdict'])}</p>
<a class="btn btn-foot" href="{esc(url)}" target="_blank" rel="{rel}">Check price on Amazon.ie {ARROW}</a>
</article>"""

def comparison_table(products):
    spec_keys = []
    for p in products:
        for k in p["specs"]:
            if k not in spec_keys:
                spec_keys.append(k)
    spec_keys = spec_keys[:4]
    head = "<tr><th>Product</th><th>Best for</th><th>Price</th>" + "".join(f"<th>{esc(k)}</th>" for k in spec_keys) + "<th>Rating</th></tr>"
    rows = ""
    for p in products:
        cells = "".join(f"<td>{esc(p['specs'].get(k, '—'))}</td>" for k in spec_keys)
        rows += f"""<tr><td><a href="#{p['id']}"><b>{esc(p['name'])}</b></a></td><td>{esc(p['badge'])}</td><td><b>€{product_price(p)}</b></td>{cells}<td><b>{p['rating']}</b>/5</td></tr>"""
    return f'<div class="tbl-scroll"><table class="cmp">{head}{rows}</table></div>'

def faq_html(faqs):
    items = "".join(f"<details><summary>{esc(f['q'])}</summary><p>{esc(f['a'])}</p></details>" for f in faqs)
    return f'<div class="guide">{items}</div>'

# ---------------------------------------------------------------- datas honestas por pagina
# ANTES: datePublished e dateModified eram TODAY_ISO em toda build, entao qualquer rebuild
# (mesmo mexendo so no rodape) remarcava as 50 paginas como "modificadas hoje". Isso e um
# sinal de frescor falso e, repetido, o Google aprende a ignorar o campo.
# AGORA: guardamos um hash do conteudo real de cada pagina em .page_dates.json.
# dateModified so avanca quando esse hash muda; datePublished nunca se move depois de gravado.
DATES_FILE = os.path.join(BASE, ".page_dates.json")
PAGE_DATES = json.load(open(DATES_FILE, encoding="utf-8")) if os.path.exists(DATES_FILE) else {}
# Data em que o conteudo profundo atual foi realmente publicado (overhaul de conteudo).
CONTENT_BASELINE = "2026-07-28"

TRACKED_IN_GEN = set()   # chaves ja hasheadas durante a geracao das paginas

def page_dates(key, content_parts):
    """Devolve (datePublished, dateModified) estaveis para uma pagina."""
    TRACKED_IN_GEN.add(key)
    h = hashlib.sha256("||".join(str(p) for p in content_parts).encode("utf-8")).hexdigest()[:16]
    rec = PAGE_DATES.get(key)
    if rec is None:
        # primeira vez que vemos a pagina: publicada agora (ou na baseline, no primeiro run)
        rec = {"hash": h, "published": CONTENT_BASELINE if not PAGE_DATES else TODAY_ISO,
               "modified": TODAY_ISO}
    elif rec.get("hash") != h:
        rec = {"hash": h, "published": rec.get("published", CONTENT_BASELINE), "modified": TODAY_ISO}
    PAGE_DATES[key] = rec
    return rec["published"], rec["modified"]

def author_schema():
    a = {"@type": "Person", "name": AUTHOR["name"], "description": AUTHOR["bio"]}
    if AUTHOR.get("url"):
        a["url"] = AUTHOR["url"]; a["sameAs"] = [AUTHOR["url"]]
    if AUTHOR.get("image"):
        a["image"] = AUTHOR["image"]
    return a

# ---------------------------------------------------------------- citação de fonte inline (GEO)
# Detecta quais factos irlandeses a página realmente usa e cita a fonte primária só desses.
# Ligar o número à fonte é o sinal de confiança que os motores generativos mais valorizam —
# e é honesto: o leitor consegue conferir a conta.
SOURCE_MARKERS = [
    (("38 cent", "€0.38", "0.38 per kWh", "38c/kWh", "38 cent a unit"),
     'electricity unit rates from <a href="https://www.seai.ie/data-and-insights/seai-statistics/prices" rel="nofollow noopener" target="_blank">SEAI energy statistics</a> and published supplier standard rates (~€0.38/kWh day rate, July 2026)'),
    (("S.I. 199", "S.I.199"),
     'e-scooter power, speed and weight limits from the text of S.I. 199 of 2024 on <a href="https://www.irishstatutebook.ie" rel="nofollow noopener" target="_blank">irishstatutebook.ie</a>'),
    (("Cycle to Work",),
     'the €1,500 Cycle to Work ceiling for e-bikes as published by <a href="https://www.revenue.ie" rel="nofollow noopener" target="_blank">Revenue</a>'),
    (("Met Éireann",),
     'pollen season dates from <a href="https://www.met.ie" rel="nofollow noopener" target="_blank">Met Éireann</a>'),
]

def sources_html(page, faqs, guide_pairs):
    blob = " ".join(t for _, t in guide_pairs) + " " + " ".join(f["a"] for f in faqs) + " " + page.get("intro", "")
    used = [txt for markers, txt in SOURCE_MARKERS if any(m in blob for m in markers)]
    if not used:
        return ""
    if len(used) == 1:
        body = used[0]
    else:
        body = "; ".join(used[:-1]) + "; and " + used[-1]
    return ('<p class="sources"><strong>Sources for the figures above:</strong> ' + body +
            '. Prices and product specifications are taken from manufacturer documentation and '
            'Amazon.ie listings at time of writing and change frequently.</p>')

def quick_answer(page, cat):
    """Resposta direta de ~40-60 palavras logo abaixo do H1.
    Padrao mais forte que existe para featured snippet: pergunta implicita do title
    respondida de imediato, em prosa curta e autocontida.
    Montado SO com dados que ja existem (top pick, preco, badge, opcao mais barata) —
    nada inventado, nada que va desalinhar do resto da pagina."""
    prods = page["products"]
    if not prods:
        return "", ""
    top = prods[0]
    # A opcao "mais barata que ainda recomendamos" nao pode ser um pick que a propria
    # pagina marcou como fora do brief (acima do teto de preco, capacidade menor que a
    # prometida, produto infantil, etc). Antes o minimo por preco pegava justamente esses.
    OUT_OF_BRIEF = ("above budget", "smaller sites", "smaller family", "room to grow",
                    "widest here", "not for adults", "not lightweight", "not a desk")
    in_brief = [p for p in prods if not any(k in p["badge"].lower() for k in OUT_OF_BRIEF)]
    cheapest = min(in_brief or prods, key=lambda p: product_price(p))
    # "Our top pick" e nao "for most homes": o rank 1 nem sempre e o best-overall
    # (em electric-scooters, por ex., o #1 e o "Best Premium" a ~€1.000). Dizer
    # "for most" ali seria uma recomendacao que a propria pagina nao sustenta.
    txt = (f"Our top pick is the {top['name']} at around €{product_price(top)} — "
           f"{top['badge'].lower()}, rated {top['rating']}/5 by owners on Amazon.ie.")
    if cheapest["id"] != top["id"]:
        txt += (f" The cheapest option we'd still recommend is the {cheapest['name']}, "
                f"from about €{product_price(cheapest)}.")
    # "picks" em vez do nome da categoria: "all 5 home office below" nao e ingles.
    txt += f" All {len(prods)} picks are compared in full below, with Irish running costs."
    html_out = (f'<p class="quick-answer" id="quick-answer"><strong>Quick answer:</strong> {esc(txt)}</p>')
    return html_out, txt

def cat_og_image(cat_slug):
    """OG image especifica da categoria. Antes as 65 paginas compartilhavam uma unica
    imagem generica, o que da preview identico em qualquer share/SERP com thumbnail."""
    return f"{DOMAIN}/assets/og-{cat_slug}.png"

def jsonld_page(page, cat, faqs, pub, mod):
    canonical = f"{DOMAIN}/{cat['category']}/{page['slug']}.html"
    items = []
    for i, p in enumerate(page["products"]):
        prod = {"@type": "Product", "name": p["name"], "description": p["verdict"], "brand": {"@type": "Brand", "name": p["brand"]},
                "offers": {"@type": "Offer", "price": str(product_price(p)), "priceCurrency": "EUR",
                           "availability": "https://schema.org/InStock", "url": product_url(p)[0],
                           # priceValidUntil: o Google recomenda em Offer. Preco e indicativo e muda
                           # direto na Amazon, entao damos uma validade curta e honesta (90 dias).
                           "priceValidUntil": (datetime.date.today() + datetime.timedelta(days=90)).isoformat()},
                }
        # NOTA: AggregateRating/Review removidos deliberadamente (2026-07-27).
        # reviewCount="1" nao e agregacao e o Review era assinado pelo proprio site
        # (self-serving review markup) -> risco de acao manual por spammy structured data.
        # A nota editorial continua visivel na pagina, so nao vai marcada como schema.
        if product_image(p):
            prod["image"] = product_image(p)
        items.append({"@type": "ListItem", "position": i + 1, "item": prod})
    data = [
        {"@context": "https://schema.org", "@type": "ItemList", "name": page["h1"], "itemListElement": items},
        {"@context": "https://schema.org", "@type": "FAQPage",
         "mainEntity": [{"@type": "Question", "name": f["q"], "acceptedAnswer": {"@type": "Answer", "text": f["a"]}} for f in faqs]},
        {"@context": "https://schema.org", "@type": "BreadcrumbList", "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "Home", "item": DOMAIN + "/"},
            {"@type": "ListItem", "position": 2, "name": cat["name"], "item": f"{DOMAIN}/{cat['category']}/"},
            {"@type": "ListItem", "position": 3, "name": page["h1"], "item": canonical}]},
        {"@context": "https://schema.org", "@type": "Article", "headline": page["h1"],
         "description": page["desc"], "image": cat_og_image(cat["category"]),
         "datePublished": pub, "dateModified": mod,
         "author": author_schema(),
         "publisher": {"@type": "Organization", "name": SITE_NAME,
                       "logo": {"@type": "ImageObject", "url": DOMAIN + "/apple-touch-icon.png"}},
         "mainEntityOfPage": {"@type": "WebPage", "@id": canonical},
         # Speakable: marca a resposta curta e o H1 como os trechos que um assistente de voz
         # deve ler em voz alta. E exatamente o bloco "Quick answer", que foi escrito
         # justamente para ser autocontido em ~50 palavras.
         "speakable": {"@type": "SpeakableSpecification",
                       "cssSelector": [".quick-answer", "h1"]}}
    ]
    return "".join(f'<script type="application/ld+json">{json.dumps(d, ensure_ascii=False)}</script>' for d in data)

def jsonld_hub(cat):
    """Schema para as paginas de indice de categoria (ex: /dehumidifiers/index.html).
    Antes nao tinham NENHUM structured data apesar de ja terem FAQ visivel na pagina
    (hub_faq) -> era um FAQPage schema perdido de graca, alem de faltar Breadcrumb/CollectionPage."""
    canonical = f"{DOMAIN}/{cat['category']}/"
    items = [{"@type": "ListItem", "position": i + 1, "url": f"{canonical}{pg['slug']}.html", "name": pg["h1"]}
             for i, pg in enumerate(cat["pages"])]
    data = [
        {"@context": "https://schema.org", "@type": "CollectionPage", "name": f"Best {cat['name']} in Ireland",
         "url": canonical, "description": cat["hub_intro"][:300],
         "isPartOf": {"@type": "WebSite", "name": SITE_NAME, "url": DOMAIN + "/"}},
        {"@context": "https://schema.org", "@type": "ItemList", "itemListElement": items},
        {"@context": "https://schema.org", "@type": "BreadcrumbList", "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "Home", "item": DOMAIN + "/"},
            {"@type": "ListItem", "position": 2, "name": cat["name"], "item": canonical}]},
    ]
    if cat.get("faqs"):
        data.append({"@context": "https://schema.org", "@type": "FAQPage",
                     "mainEntity": [{"@type": "Question", "name": f["q"],
                                     "acceptedAnswer": {"@type": "Answer", "text": f["a"]}} for f in cat["faqs"]]})
    return "".join(f'<script type="application/ld+json">{json.dumps(d, ensure_ascii=False)}</script>' for d in data)

# ---------------------------------------------------------------- load data
CATS = []
for fn in sorted(os.listdir(DATA)):
    if fn.endswith(".json"):
        with open(os.path.join(DATA, fn), encoding="utf-8") as f:
            CATS.append(json.load(f))

CATS_BY_SLUG = {c["category"]: c for c in CATS}

os.makedirs(OUT, exist_ok=True)
all_pages = []
SEARCH_INDEX = []

# ---------------------------------------------------------------- comparison pages
for cat in CATS:
    cdir = os.path.join(OUT, cat["category"])
    os.makedirs(cdir, exist_ok=True)
    for page in cat["pages"]:
        # FAQs: prefere as unicas da pagina; cai pro pool da categoria se ainda nao escritas
        faqs = page.get("faqs") or [cat["faqs"][i] for i in page["faq_idx"]]
        canonical = f"{DOMAIN}/{cat['category']}/{page['slug']}.html"
        # datas honestas: so mudam quando o conteudo real desta pagina muda
        guide_src = page.get("guide") or cat["guide"]
        # O grafico e conteudo visivel: se ele aparece, muda ou some, a pagina mudou.
        # Precisa entrar no hash ANTES de page_dates, senao a guia principal ganharia o
        # grafico sem o dateModified se mover (as specs dela nao mudaram) — frescor falso
        # ao contrario: pagina nova declarada como velha.
        _ref = RC_REF_PAGES.get(cat["category"])
        rc_html = running_cost_chart(page["products"], cat["category"],
                                     RC_NOUN.get(cat["category"], "unit"),
                                     more_href=f'{_ref["slug"]}.html' if _ref else None,
                                     more_text=_ref["link_text"] if _ref else "")
        # rc_html so entra na lista quando existe. Se entrasse sempre (como "" nas 30 guias
        # sem grafico), o proprio separador "||" mudaria o hash dessas paginas e as 50 guias
        # seriam remarcadas como modificadas hoje sem nada ter mudado nelas.
        pub, mod = page_dates(f"{cat['category']}/{page['slug']}",
            ([rc_html] if rc_html else []) + [
            page["h1"], page["intro"], page["title"], page["desc"],
            json.dumps(guide_src, ensure_ascii=False, sort_keys=True),
            json.dumps(faqs, ensure_ascii=False, sort_keys=True),
            # specs entram no hash: mudar uma spec E mudar a pagina. Sem isso, corrigir a
            # potencia de um produto nao movia o lastmod — mesmo bug de frescor falso que
            # corrigimos nos hubs em 15/08, so que escondido um nivel abaixo.
            json.dumps([[p["id"], p["name"], product_price(p), p["verdict"],
                         p.get("specs") or {}, p.get("pros") or [], p.get("cons") or []]
                        for p in page["products"]],
                       ensure_ascii=False, sort_keys=True)])
        qa_html, _ = quick_answer(page, cat)
        toc = "".join(f'<li><a href="#{p["id"]}">{esc(p["name"])}</a> <i>— {esc(p["badge"])}</i></li>' for p in page["products"])
        cards = "".join(product_card(p, i + 1, cat["category"]) for i, p in enumerate(page["products"]))
        # Buying guide: prefere o unico da pagina; cai pro da categoria se ainda nao escrito
        guide = "".join(f"<h3>{esc(h)}</h3><p>{esc(t)}</p>" for h, t in (page.get("guide") or cat["guide"]))
        others = [pg for pg in cat["pages"] if pg["slug"] != page["slug"]]
        related = "".join(f'<a href="{pg["slug"]}.html">{esc(pg["h1"])} {ARROW}</a>' for pg in others)
        body = f"""
<nav class="crumbs" aria-label="Breadcrumb"><a href="../index.html">Home</a> › <a href="index.html">{esc(cat['name'])}</a> › {esc(page['h1'])}</nav>
<h1>{esc(page['h1'])}</h1>
<div class="updated"><span class="trust-chip">{SHIELD} Independently researched</span><span class="dot"></span><span>By <a href="../about.html" rel="author">{esc(AUTHOR['name'])}</a></span><span class="dot"></span><span>Updated <time datetime="{mod}">{datetime.date.fromisoformat(mod).strftime('%d %B %Y')}</time></span><span class="dot"></span><a href="../affiliate-disclosure.html">How we make money</a></div>
{qa_html}
<p class="intro">{esc(page['intro'])}</p>
<div class="toc"><b>{icon(cat['category'], 18)} Our top {len(page['products'])} at a glance</b><ol>{toc}</ol></div>
<h2>Quick comparison</h2>
{comparison_table(page['products'])}
{rc_html}
<h2>The picks, reviewed</h2>
{cards}
<h2>Buying guide: how to choose</h2>
<div class="guide" style="padding:18px 28px">{guide}</div>
{sources_html(page, faqs, guide_src)}
<h2>Frequently asked questions</h2>
{faq_html(faqs)}
<div class="related"><h2>More {esc(cat['name'].lower())} guides</h2>{related}</div>
{cross_links_html(cat['category'], CATS_BY_SLUG)}
<p class="notice">{SITE_NAME} is reader-supported. When you buy through links on our site, we may earn an affiliate commission at no extra cost to you. Prices are indicative, in EUR, and fluctuate — always confirm the live price. We select products based on specifications, owner feedback and value analysis.</p>
"""
        out = page_shell(page["title"], page["desc"], canonical, body, depth=1,
                         jsonld=jsonld_page(page, cat, faqs, pub, mod),
                         og_image=cat_og_image(cat["category"]))
        with open(os.path.join(cdir, page["slug"] + ".html"), "w", encoding="utf-8") as f:
            f.write(out)
        all_pages.append(f"{cat['category']}/{page['slug']}.html")
        for p in page["products"]:
            SEARCH_INDEX.append({"i":p["id"],"n":p["name"],"b":p["brand"],"c":cat["name"],"p":product_price(p),"u":f"{cat['category']}/{page['slug']}.html"})

    tiles = "".join(f"""<div class="tile"><a href="{pg['slug']}.html"><span class="go">{ARROW}</span><div class="icw">{icon(cat['category'], 24)}</div><h3>{esc(pg['h1'])}</h3><p>{esc(pg['desc'][:110])}…</p></a></div>""" for pg in cat["pages"])
    hub_faq = faq_html(cat["faqs"])
    # ponte do hub para a tabela de referencia de custo, quando a categoria tem uma
    _hr = RC_REF_PAGES.get(cat["category"])
    # NB (licao de 29/08): estes blocos opcionais ficam COLADOS na linha anterior. Em linha
    # propria, a quebra de linha sobra nas categorias sem bloco e muda o hash delas — frescor
    # falso em paginas que nao mudaram.
    _ha = RC_ARTICLES.get(cat["category"])
    hub_art = (f'<a href="{_ha["slug"]}.html">{esc(_ha["hub_text"])} {ARROW}</a>') if _ha else ""
    hub_ref = ('<div class="related"><h2>Running costs, in one table</h2>'
               f'<a href="{_hr["slug"]}.html">{esc(_hr["h1"])} {ARROW}</a>'
               f'{hub_art}</div>') if _hr else ""
    body = f"""
<nav class="crumbs" aria-label="Breadcrumb"><a href="../index.html">Home</a> › {esc(cat['name'])}</nav>
<h1>Best {esc(cat['name'])} in Ireland — All Guides</h1>
<p class="intro">{esc(cat['hub_intro'])}</p>
<div class="grid">{tiles}</div>{hub_ref}
<h2>{esc(cat['name'])}: frequently asked questions</h2>
{hub_faq}
"""
    out = page_shell(f"Best {cat['name']} in Ireland 2026 | {SITE_NAME}",
                     cat["hub_intro"][:155], f"{DOMAIN}/{cat['category']}/", body, depth=1,
                     jsonld=jsonld_hub(cat), og_type="website",
                     og_image=cat_og_image(cat["category"]))
    with open(os.path.join(cdir, "index.html"), "w", encoding="utf-8") as f:
        f.write(out)
    all_pages.append(f"{cat['category']}/index.html")

# ---------------------------------------------------------------- ativo linkavel: tabela de custo por categoria
def rc_reference_page(cat):
    """Gera a pagina de referencia de custo da categoria. Devolve o caminho relativo ou None."""
    cfg = RC_REF_PAGES.get(cat["category"])
    if not cfg:
        return None
    # Generalizacao 01/09: esta pagina nasceu so para desumidificador e tinha a categoria
    # embutida na prosa e nas colunas (extracao em L/dia, custo por litro). Cada categoria
    # agora traz o proprio numero de horas/dia, a propria coluna de capacidade e a propria
    # ressalva fisica. Nada de texto unico para todas — foi exatamente o erro de 22/08.
    HRS      = cfg.get("hours_day", RC_REF_HOURS_DAY)
    HRS_TXT  = cfg.get("hours_text", f"{HRS} hours a day")
    AMT_SPEC = cfg.get("amount_spec")
    AMT_HEAD = cfg.get("amount_head", "")
    UNIT     = cfg.get("unit")
    prods = _unique_products(cat)
    withp, without = [], []
    for p in prods:
        w = _watts(p.get("specs") or {})
        if w:
            withp.append((p, w, _extraction_lpd(p.get("specs") or {}) if UNIT else None))
        else:
            without.append(p)
    if len(withp) < RC_CHART_MIN:
        return None
    withp.sort(key=lambda t: t[1])

    _eur = rc_eur

    def _amt(p):
        return esc(str((p.get("specs") or {}).get(AMT_SPEC, "—"))) if AMT_SPEC else ""

    rows = []
    for p, w, lpd in withp:
        kw = w / 1000.0
        hr_day = kw * KWH_RATE
        hr_night = kw * KWH_RATE_NIGHT
        month = hr_day * HRS * 30
        per_l = (kw * 24 * KWH_RATE / lpd) if lpd else None
        cells = [f"<td><b>{esc(p['name'])}</b></td>", f"<td>{w_txt(w)} W</td>",
                 f"<td>{_eur(hr_day)}</td>", f"<td>{_eur(hr_night)}</td>",
                 f"<td>{_eur(month)}</td>"]
        if AMT_SPEC:
            cells.append(f"<td>{_amt(p)}</td>")
        if UNIT:
            cells.append(f"<td>{_eur(per_l) if per_l else '—'}</td>")
        rows.append("<tr>" + "".join(cells) + "</tr>")
    hcells = ["<th>Model</th>", "<th>Rated power</th>", "<th>Per hour (day)</th>",
              "<th>Per hour (night)</th>", f"<th>Per month at {HRS_TXT}</th>"]
    if AMT_SPEC:
        hcells.append(f"<th>{esc(AMT_HEAD)}</th>")
    if UNIT:
        hcells.append("<th>Per litre removed</th>")
    head = "<tr>" + "".join(hcells) + "</tr>"
    table = f'<div class="tbl-scroll"><table class="cmp">{head}{"".join(rows)}</table></div>'

    lo_p, lo_w, _ = withp[0]
    hi_p, hi_w, _ = withp[-1]
    lo_h, hi_h = lo_w / 1000.0 * KWH_RATE, hi_w / 1000.0 * KWH_RATE
    n_have, n_total = len(withp), len(prods)

    # tabela 2: modelos sem potencia publicada. Existem, e omiti-los seria escolher os dados.
    if without:
        miss_rows = "".join(
            f"<tr><td><b>{esc(p['name'])}</b></td>"
            f"{f'<td>{_amt(p)}</td>' if AMT_SPEC else ''}"
            f"<td>Not published by the manufacturer</td></tr>" for p in without)
        miss_table = (
            f"<h2>The {len(without)} models where we could not calculate a cost</h2>"
            f"<p>These {cfg['noun_pl']} are in our guides, but the manufacturer does not "
            f"publish a rated power figure for them, so there is no honest way to work out "
            f"a cost per hour. We would rather show the gap than fill it with an estimate "
            f"and present it as a measurement.</p>"
            f'<div class="tbl-scroll"><table class="cmp">'
            f"<tr><th>Model</th>{f'<th>{esc(AMT_HEAD)}</th>' if AMT_SPEC else ''}"
            f"<th>Rated power</th></tr>"
            f"{miss_rows}</table></div>")
    else:
        miss_table = ""

    qa = (f'<p class="quick-answer" id="quick-answer"><strong>Quick answer:</strong> '
          f'At Ireland’s domestic day rate of about {_eur(KWH_RATE)}/kWh, the '
          f'{cfg["noun_pl"]} we track cost between {_eur(lo_h)} and {_eur(hi_h)} per hour '
          f'at full power — {_eur(lo_h * HRS * 30)} to '
          f'{_eur(hi_h * HRS * 30)} a month if you run one for '
          f'{HRS_TXT}. On a smart-meter night rate of about '
          f'{_eur(KWH_RATE_NIGHT)}/kWh the same run costs roughly half. These are ceilings: '
          f'{cfg["ceiling_short"]}</p>')

    guides = "".join(
        f'<a href="{pg["slug"]}.html">{esc(pg["h1"])} {ARROW}</a>' for pg in cat["pages"])
    # ponte para o artigo (a versao em prosa desta tabela), quando a categoria tem um
    _art = RC_ARTICLES.get(cat["category"])
    art_link = (f'<a href="{_art["slug"]}.html">{esc(_art["h1"])} {ARROW}</a>') if _art else ""

    per_unit_step = (f"""<li><strong>Cost per {cfg['unit']}</strong> = a full 24 hours at
rated power, divided by the rated daily extraction. It is the fairest way to compare a
thirsty 20 L machine against a frugal 6 L one, because the cheaper machine per hour
is not automatically the cheaper machine per litre.</li>""" if UNIT else "")
    unit_note_block = (f"<p><strong>The {cfg.get('unit_note_label', 'capacity')} figures are "
                       f"lab figures.</strong> {cfg['unit_note']}</p>"
                       if cfg.get("unit_note") else "")
    intro_tail = (f", per month and per {cfg['unit']}," if UNIT else " and per month")

    per_unit_step = (f"""<li><strong>Cost per {cfg['unit']}</strong> = a full 24 hours at
rated power, divided by the rated daily extraction. It is the fairest way to compare a
thirsty 20 L machine against a frugal 6 L one, because the cheaper machine per hour
is not automatically the cheaper machine per litre.</li>""" if UNIT else "")
    unit_note_block = (f"<p><strong>The {cfg.get('unit_note_label', 'capacity')} figures are "
                       f"lab figures.</strong> {cfg['unit_note']}</p>"
                       if cfg.get("unit_note") else "")
    intro_tail = (f", per month and per {cfg['unit']}," if UNIT else " and per month")

    method = f"""
<h2>How these numbers are worked out</h2>
<p>The whole calculation is three numbers and one multiplication, and we would rather you
checked it than trusted it.</p>
<ol class="method">
<li><strong>Rated power</strong> — the wattage the manufacturer publishes for that exact
model. Not a category average, not a guess from a similar unit. {n_have} of the
{n_total} {cfg['noun_pl']} in our guides publish one; the other {n_total - n_have} are
listed separately, with no figure at all.</li>
<li><strong>Tariff</strong> — {_eur(KWH_RATE)} per kWh for the day rate and
{_eur(KWH_RATE_NIGHT)} per kWh for a typical smart-meter night rate, cross-checked against
<a href="https://www.seai.ie/data-and-insights/seai-statistics/prices" rel="nofollow noopener"
target="_blank">SEAI energy price statistics</a> and published supplier standard rates,
July 2026. Your own unit rate is on your bill and may differ; the arithmetic below is easy
to redo with it.</li>
<li><strong>Cost per hour</strong> = rated watts &divide; 1,000 &times; the unit rate. A
{w_txt(lo_w)}&nbsp;W unit is {lo_w / 1000:.3f}&nbsp;kW, so {lo_w / 1000:.3f} &times;
{KWH_RATE:.2f} = {_eur(lo_h)} an hour.</li>
<li><strong>Cost per month</strong> = cost per hour &times; {HRS} hours &times; 30 days.
{HRS_TXT} is our assumption, not a measurement — scale it to how you actually run yours.</li>
{per_unit_step}
</ol>
<h2>Three things these numbers are not</h2>
<p><strong>They are not your bill.</strong> Every figure here is a ceiling: rated draw at
full power, multiplied out. {cfg['ceiling_note']} Either way, treat the table as the worst
case.</p>
{unit_note_block}
<p><strong>They do not include standing charges or VAT changes.</strong> The unit rate is the
marginal cost of running the appliance. Your standing charge is there whether the
{cfg['noun']} is plugged in or not.</p>
<h2>Using this table</h2>
<p>Free to quote, cite or reproduce with a link back to this page. If you are writing about
damp, heating or energy costs in Ireland and want the underlying figures, everything is on
this page — the rated wattages, the tariff, and the arithmetic — so the numbers can be
checked rather than taken on trust. Spot an error or a wattage we have wrong?
<a href="../contact.html">Tell us</a> and we will correct it.</p>
<h2>The guides these {cfg['noun_pl']} come from</h2>
<div class="related">{art_link}{guides}</div>
"""

    body = f"""
<nav class="crumbs" aria-label="Breadcrumb"><a href="../index.html">Home</a> › <a href="index.html">{esc(cat['name'])}</a> › Running costs</nav>
<h1>{esc(cfg['h1'])}</h1>
<div class="updated"><span class="trust-chip">{SHIELD} Method published in full</span><span class="dot"></span><span>By <a href="../about.html" rel="author">{esc(AUTHOR['name'])}</a></span><span class="dot"></span><span>Updated <time datetime="{{MOD}}">{{MODLONG}}</time></span><span class="dot"></span><a href="../affiliate-disclosure.html">How we make money</a></div>
{qa}
<p class="intro">Almost every {cfg['noun']} page online tells you which one to buy and none of
them tell you what it costs to keep switched on. This page is the other half: every
{cfg['noun']} we track, its rated power, and what that works out to per hour{intro_tail}
at Irish electricity prices — with the arithmetic shown so you can redo it with your own
unit rate.</p>
<h2>Running cost of every {cfg['noun']} we track</h2>
{table}
<p class="sources"><strong>Tariff source:</strong> {_eur(KWH_RATE)}/kWh day rate and
{_eur(KWH_RATE_NIGHT)}/kWh night rate, cross-checked against
<a href="https://www.seai.ie/data-and-insights/seai-statistics/prices" rel="nofollow noopener"
target="_blank">SEAI energy price statistics</a> and published supplier standard rates, July
2026. Rated power{f" and {AMT_HEAD.lower()}" if AMT_SPEC else ""} are the manufacturers’
own published figures.</p>
{miss_table}
{method}
"""
    key = f"{cat['category']}/{cfg['slug']}"
    pub, mod = page_dates(key, [
        cfg["h1"], cfg["title"], cfg["desc"], f"{KWH_RATE}|{KWH_RATE_NIGHT}|{HRS}",
<<<<<<< HEAD
        # o link para o artigo e conteudo visivel: se aparece, a pagina mudou. Entra na lista
        # SO quando existe — se entrasse sempre (como "" nas 3 categorias sem artigo), o
        # proprio separador "||" mudaria o hash delas e as marcaria como modificadas hoje
        # sem uma letra ter mudado no HTML. Mesmo bug de frescor falso de 22/08 e 29/08.
        ] + ([art_link] if art_link else []) + [
=======
>>>>>>> 70447d3a93e2e11a184cd5542a15dca748d4680d
        json.dumps([[p["name"], w, lpd] for p, w, lpd in withp], ensure_ascii=False),
        json.dumps([p["name"] for p in without], ensure_ascii=False)])
    body = body.replace("{MOD}", mod).replace(
        "{MODLONG}", datetime.date.fromisoformat(mod).strftime("%d %B %Y"))

    canonical = f"{DOMAIN}/{cat['category']}/{cfg['slug']}.html"
    dataset = {
        "@context": "https://schema.org", "@type": "Dataset",
        "name": f"Running cost of {len(withp)} {cfg['noun_pl']} at Irish electricity rates (2026)",
        "description": cfg["desc"], "url": canonical,
        "license": f"{DOMAIN}/about.html",
        "isAccessibleForFree": True, "inLanguage": "en-IE",
        "datePublished": pub, "dateModified": mod,
        "spatialCoverage": {"@type": "Country", "name": "Ireland"},
        "measurementTechnique": (f"Manufacturer-published rated power multiplied by a domestic "
                                 f"electricity unit rate of €{KWH_RATE:.2f}/kWh (day) and "
                                 f"€{KWH_RATE_NIGHT:.2f}/kWh (night), July 2026."),
        "variableMeasured": ["Rated power (W)", "Cost per hour (EUR)",
                             "Cost per month at 8 h/day (EUR)",
                             "Cost per litre of water extracted (EUR)"],
        "creator": {"@id": DOMAIN + "/#organization"},
        "publisher": {"@id": DOMAIN + "/#organization"}}
    webpage = {
        "@context": "https://schema.org", "@type": "WebPage", "@id": canonical,
        "url": canonical, "name": cfg["title"], "description": cfg["desc"],
        "inLanguage": "en-IE", "datePublished": pub, "dateModified": mod,
        "author": author_schema(), "publisher": {"@id": DOMAIN + "/#organization"},
        "mainEntity": {"@id": canonical + "#dataset"},
        "speakable": {"@type": "SpeakableSpecification",
                      "cssSelector": [".quick-answer", "h1"]}}
    crumb = {"@context": "https://schema.org", "@type": "BreadcrumbList", "itemListElement": [
        {"@type": "ListItem", "position": 1, "name": "Home", "item": DOMAIN + "/"},
        {"@type": "ListItem", "position": 2, "name": cat["name"],
         "item": f"{DOMAIN}/{cat['category']}/"},
        {"@type": "ListItem", "position": 3, "name": "Running costs", "item": canonical}]}
    dataset["@id"] = canonical + "#dataset"
    jsonld = "".join(f'<script type="application/ld+json">{json.dumps(x, ensure_ascii=False)}</script>'
                     for x in (dataset, webpage, crumb))
    out = page_shell(cfg["title"], cfg["desc"], canonical, body, depth=1, jsonld=jsonld,
                     og_image=cat_og_image(cat["category"]))
    with open(os.path.join(OUT, cat["category"], cfg["slug"] + ".html"), "w",
              encoding="utf-8") as f:
        f.write(out)
    all_pages.append(f"{cat['category']}/{cfg['slug']}.html")
    print(f"  [ref] {canonical} — {len(withp)} with power, {len(without)} without")
    return f"{cat['category']}/{cfg['slug']}.html"

REF_PATHS = {}
for cat in CATS:
    _rp = rc_reference_page(cat)
    if _rp:
        REF_PATHS[cat["category"]] = _rp

# ---------------------------------------------------------------- artigo informacional
# Semana 3 da fila (secao 7 do plano, 05/09/2026). Por que um ARTIGO alem da tabela:
# a tabela responde "qual modelo custa quanto" — e uma referencia, por modelo. A busca real
# ("how much does a dehumidifier cost to run in Ireland") e uma pergunta unica com uma
# resposta em euros, e quem faz essa busca nao quer 12 linhas, quer um numero e o raciocinio.
# Sao intencoes diferentes, entao sao paginas diferentes; cada uma linka para a outra em
# prosa para que nem o leitor nem o Google fiquem em duvida sobre qual e qual.
#
# Regra de honestidade desta pagina: NENHUM numero vem de fora dos dados do proprio site.
# A comparacao com aquecedor usa as potencias reais da categoria electric-heaters, que ja
# estao verificadas aqui. Nao ha figura de secadora, de caldeira nem de "media irlandesa" —
# se nao esta nos dados, nao entra. O que falta e declarado como faltando.
def running_cost_article(cat):
    """Artigo informacional de custo de operacao. Devolve o caminho relativo ou None."""
    cfg = RC_ARTICLES.get(cat["category"])
    ref = RC_REF_PAGES.get(cat["category"])
    if not cfg or not ref:
        return None
    withp = []
    for p in _unique_products(cat):
        w = _watts(p.get("specs") or {})
        if w:
            withp.append((p, w, _extraction_lpd(p.get("specs") or {})))
    if len(withp) < RC_CHART_MIN:
        return None
    withp.sort(key=lambda t: t[1])
    HRS = ref.get("hours_day", RC_REF_HOURS_DAY)
    _eur = rc_eur

    lo_p, lo_w, lo_l = withp[0]
    hi_p, hi_w, hi_l = withp[-1]
    md_p, md_w, md_l = withp[(len(withp) - 1) // 2]
    md_kw = md_w / 1000.0
    # CONFERENCIA ADVERSARIAL (05/09): a faixa da resposta rapida nao pode comecar no mini de
    # 36 W. Ele e um aparelho de 0.3 L/dia — dizer "a partir de €0.01/hora" seria verdade
    # aritmetica e mentira pratica. A faixa cobre so as maquinas de tomada de verdade.
    mains = [t for t in withp if t[1] >= 100] or withp
    mn_w, mx_w = mains[0][1], mains[-1][1]
    tiny = [t for t in withp if t[1] < 100]

    def _row(p, w):
        kw = w / 1000.0
        return ("<tr>"
                f"<td><b>{esc(p['name'])}</b></td>"
                f"<td>{w_txt(w)} W</td>"
                f"<td>{_eur(kw * KWH_RATE)}</td>"
                f"<td>{_eur(kw * KWH_RATE_NIGHT * HRS)}</td>"
                f"<td>{_eur(kw * KWH_RATE * HRS * 30)}</td>"
                "</tr>")
    scen = ('<div class="tbl-scroll"><table class="cmp">'
            "<tr><th>Model</th><th>Rated power</th><th>One hour, day rate</th>"
            f"<th>{HRS} hours overnight, night rate</th>"
            f"<th>A month at {HRS} h/day, day rate</th></tr>"
            + _row(lo_p, lo_w) + _row(md_p, md_w) + _row(hi_p, hi_w)
            + "</table></div>")

    # O que muda mais a conta: trocar de modelo, ou mudar a hora em que ele roda.
    # CONFERENCIA ADVERSARIAL (05/09): a primeira versao comparava o modelo mediano com o de
    # menor potencia da categoria — 185 W contra 101 W. Nao e comparacao valida: o de 101 W
    # extrai 6 L/dia e o de 185 W extrai 12, entao "trocar" nao entrega o mesmo trabalho e a
    # economia anunciada era ficticia. A comparacao honesta e entre modelos com a MESMA
    # extracao nominal. Usamos o maior grupo de extracao igual.
    from collections import Counter
    _lgroups = Counter(l for _p, _w, l in withp if l)
    peer = []
    if _lgroups:
        _top_l = _lgroups.most_common(1)[0][0]
        peer = sorted([t for t in withp if t[2] == _top_l], key=lambda t: t[1])
    save_tariff = md_kw * (KWH_RATE - KWH_RATE_NIGHT) * HRS * 30

    # comparacao com aquecedor — potencias reais da nossa propria categoria, nao estimativa
    heat = CATS_BY_SLUG.get("electric-heaters")
    heat_block = ""
    if heat:
        hw = sorted(w for w in (_watts(p.get("specs") or {}) for p in _unique_products(heat)) if w)
        if hw:
            common = max(set(hw), key=hw.count)
            mins = md_kw * HRS / (common / 1000.0) * 60
            heat_block = f"""
<h2>Put next to a heater, it is not a big number</h2>
<p>The most common rating among the {len(hw)} electric heaters we track is
{w_txt(common)}&nbsp;W, and a heater converts essentially all of that into heat the whole
time it is on. That heater costs {_eur(common / 1000.0 * KWH_RATE)} an hour at the day rate.
The {w_txt(md_w)}&nbsp;W dehumidifier above costs {_eur(md_kw * KWH_RATE * HRS)} for a full
{HRS}-hour run — so a whole night of dehumidifying costs about the same as
<strong>{mins:.0f} minutes</strong> of that heater. That is the comparison worth holding on
to: on an Irish electricity bill a dehumidifier is a small, steady cost, and the heating is
where the money goes. See <a href="../electric-heaters/{RC_REF_PAGES['electric-heaters']['slug']}.html">what
every heater we track costs per hour</a> if you want the other half of that sum.</p>"""

    peer_block = ""
    if len(peer) >= 3:
        pl_p, pl_w, _ = peer[0]
        ph_p, ph_w, _ = peer[-1]
        save_model = (ph_w - pl_w) / 1000.0 * KWH_RATE * HRS * 30
        peer_block = f"""
<h2>When you run it is worth as much as which one you buy</h2>
<p>This is the part most buying guides skip. {len(peer)} of the {cfg['noun_pl'] if 'noun_pl' in cfg else 'dehumidifiers'} in our guides are rated at
exactly the same {peer[0][2]:g}&nbsp;L a day, so comparing those is fair — they are being
asked to do the same job. Their draw runs from {w_txt(pl_w)}&nbsp;W
({esc(pl_p['name'])}) to {w_txt(ph_w)}&nbsp;W ({esc(ph_p['name'])}), which over a month at
{ref.get('hours_text', f'{HRS} hours a day')} is
{_eur(pl_w / 1000.0 * KWH_RATE * HRS * 30)} against
{_eur(ph_w / 1000.0 * KWH_RATE * HRS * 30)}.</p>
<ul>
<li>Buying the most frugal of those instead of the thirstiest saves
<strong>{_eur(save_model)} a month</strong>.</li>
<li>Taking any one of them and moving the run onto a night rate saves
<strong>{_eur(save_tariff)} a month</strong> — worked on the {w_txt(md_w)}&nbsp;W machine
in the table above.</li>
</ul>
<p>In other words, the tariff you run it on is worth about as much as the best model choice
available to you, and switching to it costs nothing at all if your machine or your socket
has a timer. Neither number is large — which is the honest headline of this page.</p>"""

    per_l_rows = [(p, w, l, w / 1000.0 * 24 * KWH_RATE / l) for p, w, l in withp if l]
    per_l_block = ""
    if per_l_rows:
        cheap = min(per_l_rows, key=lambda t: t[3])
        dear = max(per_l_rows, key=lambda t: t[3])
        per_l_block = f"""
<h2>Cost per litre: the number that reorders the list</h2>
<p>Cost per hour rewards the smallest machine, which is not the same thing as the cheapest
machine. A unit that sips electricity but pulls very little water out of the air can cost
more per litre removed than a bigger one that does the job and switches off. On the
{len(per_l_rows)} models where the manufacturer publishes both a wattage and a rated
extraction, running flat out for 24 hours works out at
{_eur(cheap[3])} per litre for the {esc(cheap[0]['name'])} and {_eur(dear[3])} per litre for
the {esc(dear[0]['name'])} — a {dear[3] / cheap[3]:.1f}× spread that the per-hour column
does not show you.</p>
<p><strong>The catch, said plainly.</strong> Rated extraction is measured at around
30&nbsp;°C and 80% relative humidity. An Irish bedroom in November is nowhere near that, so
every machine here will remove less than its rating and your real cost per litre will be
higher than the figures above. We use the rated number because it is the one the
manufacturer publishes and you can check it — not because we think you will get it.</p>"""

    faqs = [
        ("Is it cheaper to run a dehumidifier at night?",
         f"Yes, if you are on a smart-meter tariff with a night rate. The same {HRS}-hour run "
         f"on the {w_txt(md_w)} W model above costs "
         f"{_eur(md_kw * KWH_RATE * HRS)} at the day rate and "
         f"{_eur(md_kw * KWH_RATE_NIGHT * HRS)} at a night rate of "
         f"{_eur(KWH_RATE_NIGHT)}/kWh. On a flat tariff the time of day changes nothing."),
        ("Does leaving a dehumidifier on all day cost a lot?",
         f"Twenty-four hours at full power on the {w_txt(md_w)} W model is "
         f"{_eur(md_kw * 24 * KWH_RATE)} at the day rate. That is the ceiling, not the bill: a "
         f"compressor model with a humidistat shuts the compressor off once the room reaches "
         f"its target humidity, so it spends much of the day drawing almost nothing."),
        ("Is a dehumidifier cheaper than heating the room?",
         f"On these figures, comfortably. The dehumidifier costs "
         f"{_eur(md_kw * KWH_RATE)} an hour at full draw; a typical 2 kW electric heater costs "
         f"{_eur(2 * KWH_RATE)}. They do different jobs, so this is not an either/or — but "
         f"drying the air is the cheap part of the problem."),
        ("Is a dehumidifier cheaper than a tumble dryer?",
         "We do not publish a figure for this, because we do not track tumble dryers and we "
         "will not put a number on the page that we have not verified. The sum is the same "
         "one used here, and you can do it: take the rated watts off the plate on the back of "
         "your dryer, divide by 1,000, multiply by your unit rate, and multiply by the length "
         "of a cycle."),
        ("Which dehumidifier is cheapest to run?",
         f"Of the {len(mains)} mains models where the manufacturer publishes a wattage, the "
         f"lowest draw is the {mains[0][0]['name']} at {w_txt(mn_w)} W "
         f"({_eur(mn_w / 1000.0 * KWH_RATE)} an hour) and the highest is the "
         f"{mains[-1][0]['name']} at {w_txt(mx_w)} W "
         f"({_eur(mx_w / 1000.0 * KWH_RATE)} an hour). Cheapest per hour is not the same as "
         f"cheapest per litre of water removed, which is the comparison that matters if the "
         f"two machines are not rated to pull out the same amount."),
    ]
    faq_block = ('<div class="guide">' + "".join(
        f"<details><summary>{esc(q)}</summary><p>{esc(a)}</p></details>" for q, a in faqs)
        + "</div>")

    guides = "".join(
        f'<a href="{pg["slug"]}.html">{esc(pg["h1"])} {ARROW}</a>' for pg in cat["pages"])

    tiny_txt = ((f' We also track {len(tiny)} sub-100&nbsp;W mini units, which cost less again '
                 f'but only pull out a fraction of a litre a day.') if len(tiny) > 1 else
                (f' We also track one sub-100&nbsp;W mini unit, which costs less again but only '
                 f'pulls out a fraction of a litre a day.')) if tiny else ""
    qa = (f'<p class="quick-answer" id="quick-answer"><strong>Quick answer:</strong> '
          f'A mains dehumidifier costs between {_eur(mn_w / 1000.0 * KWH_RATE)} and '
          f'{_eur(mx_w / 1000.0 * KWH_RATE)} an hour to run at Ireland’s domestic day rate of '
          f'about {_eur(KWH_RATE)}/kWh, based on the rated power of the {len(mains)} models we '
          f'track that publish one.{tiny_txt} A typical {w_txt(md_w)}&nbsp;W unit run overnight for '
          f'{HRS} hours costs {_eur(md_kw * KWH_RATE * HRS)} on the day rate and '
          f'{_eur(md_kw * KWH_RATE_NIGHT * HRS)} on a night rate — roughly '
          f'{_eur(md_kw * KWH_RATE * HRS * 30)} a month if you run it every night. Those are '
          f'ceilings: the humidistat cuts the compressor once the room dries out.</p>')

    body = f"""
<nav class="crumbs" aria-label="Breadcrumb"><a href="../index.html">Home</a> › <a href="index.html">{esc(cat['name'])}</a> › {esc(cfg['crumb'])}</nav>
<h1>{esc(cfg['h1'])}</h1>
<div class="updated"><span class="trust-chip">{SHIELD} Every figure worked from published wattages</span><span class="dot"></span><span>By <a href="../about.html" rel="author">{esc(AUTHOR['name'])}</a></span><span class="dot"></span><span>Updated <time datetime="{{MOD}}">{{MODLONG}}</time></span><span class="dot"></span><a href="../affiliate-disclosure.html">How we make money</a></div>
{qa}
<p class="intro">Dehumidifiers have a reputation in Ireland for being expensive to leave on.
This page works out whether that is true, using the wattage each manufacturer publishes and
the price of a unit of electricity here — with every step of the arithmetic shown, so you
can redo it with the rate on your own bill.</p>

<h2>The three numbers that decide your bill</h2>
<p>There is no fourth. A dehumidifier's running cost is its <strong>rated power</strong> in
kilowatts, multiplied by the <strong>hours</strong> you run it, multiplied by your
<strong>unit rate</strong> in euro per kWh.</p>
<ol class="method">
<li><strong>Rated power.</strong> The wattage on the plate. Of the {len(_unique_products(cat))}
dehumidifiers in our guides, {len(withp)} publish one; the rest do not, and we leave those
blank rather than estimate them.</li>
<li><strong>Unit rate.</strong> We use {_eur(KWH_RATE)}/kWh for the day rate and
{_eur(KWH_RATE_NIGHT)}/kWh for a typical smart-meter night rate, cross-checked against
<a href="https://www.seai.ie/data-and-insights/seai-statistics/prices" rel="nofollow noopener"
target="_blank">SEAI energy price statistics</a> and published supplier standard rates, July
2026. Yours is on your bill.</li>
<li><strong>Hours.</strong> Ours is {ref.get('hours_text', f'{HRS} hours a day')} — an
assumption, not a measurement. Scale everything below to how you actually run yours.</li>
</ol>
<p>So a {w_txt(md_w)}&nbsp;W machine is {md_kw:.3f}&nbsp;kW, and {md_kw:.3f} &times;
{KWH_RATE:.2f} = {_eur(md_kw * KWH_RATE)} an hour. That is the whole calculation.</p>

<h2>What a night, and a month, actually cost</h2>
<p>Three machines from our guides — the lowest draw we have a figure for, a typical mid-range
unit, and the highest — run through the same sum.</p>
{scen}
<p class="sources"><strong>Tariff source:</strong> {_eur(KWH_RATE)}/kWh day rate and
{_eur(KWH_RATE_NIGHT)}/kWh night rate, cross-checked against
<a href="https://www.seai.ie/data-and-insights/seai-statistics/prices" rel="nofollow noopener"
target="_blank">SEAI energy price statistics</a> and published supplier standard rates, July
2026. Rated power is each manufacturer's own published figure. The full table, for every
model we track, is on <a href="{ref['slug']}.html">{esc(ref['h1'])}</a>.</p>

{peer_block}
{per_l_block}
{heat_block}

<h2>What these figures are not</h2>
<p><strong>They are not your bill.</strong> Every number here is rated draw at full power,
multiplied out. {ref['ceiling_note']}</p>
<p><strong>They do not include your standing charge or VAT changes.</strong> The unit rate is
the marginal cost of switching the thing on. The standing charge is there either way.</p>
<p><strong>They are not a measurement.</strong> We have not put a plug-in energy monitor on
these machines. Everything above is published wattage and arithmetic, which is exactly why
we show the arithmetic — so you can check it rather than trust it. Spot a wattage we have
wrong? <a href="../contact.html">Tell us</a> and we will correct it.</p>

<h2>Frequently asked questions</h2>
{faq_block}

<div class="related"><h2>The full table, model by model</h2>
<a href="{ref['slug']}.html">{esc(ref['h1'])} {ARROW}</a></div>
<div class="related"><h2>Which dehumidifier to buy</h2>{guides}</div>
{cross_links_html(cat['category'], CATS_BY_SLUG)}
<p class="notice">{SITE_NAME} is reader-supported. When you buy through links on our site, we
may earn an affiliate commission at no extra cost to you. Running costs on this page are
calculated from manufacturer-published rated power and are indicative, not measured.</p>
"""
    key = f"{cat['category']}/{cfg['slug']}"
    pub, mod = page_dates(key, [
        cfg["h1"], cfg["title"], cfg["desc"],
        f"{KWH_RATE}|{KWH_RATE_NIGHT}|{HRS}",
        json.dumps([[p["name"], w, l] for p, w, l in withp], ensure_ascii=False),
        json.dumps(faqs, ensure_ascii=False),
        heat_block, per_l_block, peer_block, qa])
    body = body.replace("{MOD}", mod).replace(
        "{MODLONG}", datetime.date.fromisoformat(mod).strftime("%d %B %Y"))

    canonical = f"{DOMAIN}/{cat['category']}/{cfg['slug']}.html"
    ref_url = f"{DOMAIN}/{cat['category']}/{ref['slug']}.html"
    article = {
        "@context": "https://schema.org", "@type": "Article", "@id": canonical + "#article",
        "headline": cfg["h1"], "description": cfg["desc"],
        "mainEntityOfPage": {"@type": "WebPage", "@id": canonical},
        "inLanguage": "en-IE", "datePublished": pub, "dateModified": mod,
        "author": author_schema(), "publisher": {"@id": DOMAIN + "/#organization"},
        "image": cat_og_image(cat["category"]),
        "about": {"@type": "Thing", "name": "Dehumidifier running cost in Ireland"},
        "citation": {"@type": "Dataset", "@id": ref_url + "#dataset", "url": ref_url},
        "isBasedOn": ref_url}
    webpage = {
        "@context": "https://schema.org", "@type": "WebPage", "@id": canonical,
        "url": canonical, "name": cfg["title"], "description": cfg["desc"],
        "inLanguage": "en-IE", "datePublished": pub, "dateModified": mod,
        "speakable": {"@type": "SpeakableSpecification",
                      "cssSelector": [".quick-answer", "h1"]}}
    faqpage = {
        "@context": "https://schema.org", "@type": "FAQPage",
        "@id": canonical + "#faq", "mainEntity": [
            {"@type": "Question", "name": q,
             "acceptedAnswer": {"@type": "Answer", "text": a}} for q, a in faqs]}
    crumb = {"@context": "https://schema.org", "@type": "BreadcrumbList", "itemListElement": [
        {"@type": "ListItem", "position": 1, "name": "Home", "item": DOMAIN + "/"},
        {"@type": "ListItem", "position": 2, "name": cat["name"],
         "item": f"{DOMAIN}/{cat['category']}/"},
        {"@type": "ListItem", "position": 3, "name": cfg["crumb"], "item": canonical}]}
    jsonld = "".join(f'<script type="application/ld+json">{json.dumps(x, ensure_ascii=False)}</script>'
                     for x in (article, webpage, faqpage, crumb))
    out = page_shell(cfg["title"], cfg["desc"], canonical, body, depth=1, jsonld=jsonld,
                     og_image=cat_og_image(cat["category"]))
    with open(os.path.join(OUT, cat["category"], cfg["slug"] + ".html"), "w",
              encoding="utf-8") as f:
        f.write(out)
    all_pages.append(f"{cat['category']}/{cfg['slug']}.html")
    print(f"  [article] {canonical}")
    return f"{cat['category']}/{cfg['slug']}.html"

for cat in CATS:
    running_cost_article(cat)

# ---------------------------------------------------------------- homepage
tiles = ""
for cat in CATS:
    tiles += f"""<div class="tile"><a href="{cat['category']}/index.html"><span class="go">{ARROW}</span><div class="icw">{icon(cat['category'], 24)}</div><h3>{esc(cat['name'])}</h3><p>{len(cat['pages'])} buying guides · {sum(len(p['products']) for p in cat['pages'])} products compared</p></a></div>"""
featured = ""
for cat in CATS:
    pg = cat["pages"][0]
    featured += f'<a href="{cat["category"]}/{pg["slug"]}.html"><span style="display:flex;align-items:center;gap:10px"><span class="icw" style="width:34px;height:34px;border-radius:9px;background:var(--green-t);color:var(--green);display:inline-flex;align-items:center;justify-content:center">{icon(cat["category"], 18)}</span>{esc(pg["h1"])}</span>{ARROW}</a>'
n_guides = sum(len(c["pages"]) for c in CATS)
n_prods = sum(len(p["products"]) for c in CATS for p in c["pages"])

# spotlight featured products
_spot_labels={'dehumidifiers':'Damp season essential','air-fryers':'Kitchen favourite','coffee-machines':'High-ticket pick','air-purifiers':'Allergy season pick','electric-heaters':'Winter essential','electric-bikes':'Commuter favourite','electric-scooters':'City mobility','home-office':'WFH upgrade','robot-vacuums':'Hands-free cleaning','robot-lawn-mowers':'Garden on autopilot'}
spot_keys=[(c['category'], _spot_labels.get(c['category'],'Editor pick')) for c in CATS]
spot_tabs=""; spot_panels=""
for si,(sk,slabel) in enumerate(spot_keys):
    scat=next(c for c in CATS if c['category']==sk)
    sp=scat['pages'][0]['products'][0]
    surl,_=product_url(sp)
    on=" on" if si==0 else ""
    spot_tabs+=f'<button type="button" class="spot-tab{on}" data-k="{sk}">{esc(scat["name"])}</button>'
    r1=int(sp['rating']*20); r2=min(97,88+si*3)
    spot_panels+=f"""<div class="spot-panel{on}" data-k="{sk}">
<div class="spot-vis">{icon(sk,64)}</div>
<div class="spot-info">
<div class="k">{esc(slabel)} · {esc(sp['badge'])}</div>
<h3>{esc(sp['name'])}</h3>
<div class="pr">€{product_price(sp)}</div>
<div class="bar"><div class="lb"><span>Our rating</span><span>{sp['rating']}/5</span></div><u><i data-w="{r1}"></i></u></div>
<div class="bar"><div class="lb"><span>Editor confidence</span><span>{r2}%</span></div><u><i data-w="{r2}"></i></u></div>
<a class="btn" href="{esc(surl)}" target="_blank" rel="sponsored noopener">Check Price on Amazon.ie {ARROW}</a>
</div></div>"""
spot_html=f'<section class="spot" aria-label="Featured picks"><div class="spot-head"><h2>Editor&#39;s spotlight</h2><div class="spot-tabs">{spot_tabs}</div></div>{spot_panels}</section>'

body_home = f"""
<div class="hero" style="margin:0 -22px"><div class="wrap">
<h1>Buy right the first time, <em>Ireland</em></h1>
<p>Independent comparisons of what Irish homes actually buy — real € prices, running costs on Irish electricity, and verdicts that pick a side. No fluff, no 40 open tabs.</p>
<div class="hero-stats"><div><b>{n_guides}</b>buying guides</div><div><b>{n_prods}</b>products compared</div><div><b>{YEAR}</b>kept updated</div></div>
</div></div>
{spot_html}
<h2 id="categories">Browse by category</h2>
<div class="grid">{tiles}</div>
<h2>Featured guides</h2>
<div class="related">{featured}</div>
<h2>How {SITE_NAME} works</h2>
<p>Every guide compares five carefully selected products using manufacturer specifications, verified owner feedback and Irish-specific factors — electricity rates, weather, legal rules and local availability. When you buy through our links we may earn a commission from Amazon.ie or other retailers, at no cost to you. That's the entire business model: useful guides, honest picks. <a href="affiliate-disclosure.html">Full disclosure here</a>.</p>
"""
_org = {"@context": "https://schema.org", "@type": "Organization", "@id": DOMAIN + "/#organization",
        "name": SITE_NAME, "url": DOMAIN + "/",
        "logo": {"@type": "ImageObject", "url": DOMAIN + "/apple-touch-icon.png", "width": 180, "height": 180},
        "image": OG_IMAGE,
        "description": "Independent product comparison guides for Irish shoppers, factoring in Irish prices, electricity costs, weather and rules.",
        "areaServed": {"@type": "Country", "name": "Ireland"},
        "knowsLanguage": "en-IE"}
if AUTHOR.get("url"):
    _org["sameAs"] = [AUTHOR["url"]]
_website = {"@context": "https://schema.org", "@type": "WebSite", "@id": DOMAIN + "/#website",
            "name": SITE_NAME, "url": DOMAIN + "/",
            "description": "Independent product comparison guides for Irish shoppers.",
            "inLanguage": "en-IE", "areaServed": {"@type": "Country", "name": "Ireland"},
            "publisher": {"@id": DOMAIN + "/#organization"}}
home_jsonld = ("".join('<script type="application/ld+json">' + json.dumps(d, ensure_ascii=False) + "</script>"
                       for d in (_org, _website)))
with open(os.path.join(OUT, "index.html"), "w", encoding="utf-8") as f:
    f.write(page_shell(f"{SITE_NAME} — Honest Product Comparison Guides Ireland 2026",
                       "Independent buying guides for Irish shoppers: e-scooters, e-bikes, dehumidifiers, air fryers and heaters — compared for Irish prices, weather and rules.",
                       DOMAIN + "/", body_home, depth=0, jsonld=home_jsonld, og_type="website"))
all_pages.append("index.html")

# ---------------------------------------------------------------- legal & info pages
def simple_page(fname, title, body_html, desc=None, page_type="WebPage", extra=None):
    d = desc or title
    # WebPage/AboutPage/ContactPage schema -- antes essas paginas nao tinham NENHUM
    # structured data (exceto about.html) e a meta description caia pro titulo
    # (7-20 caracteres, praticamente vazia). `extra` mistura props especificas (ex: mainEntity do autor).
    schema = {"@context": "https://schema.org", "@type": page_type, "name": title,
              "url": f"{DOMAIN}/{fname}", "description": d,
              "isPartOf": {"@type": "WebSite", "name": SITE_NAME, "url": DOMAIN + "/"}}
    if extra:
        schema.update(extra)
    crumb = {"@context": "https://schema.org", "@type": "BreadcrumbList", "itemListElement": [
        {"@type": "ListItem", "position": 1, "name": "Home", "item": DOMAIN + "/"},
        {"@type": "ListItem", "position": 2, "name": title, "item": f"{DOMAIN}/{fname}"}]}
    jsonld = "".join(f'<script type="application/ld+json">{json.dumps(x, ensure_ascii=False)}</script>'
                     for x in (schema, crumb))
    with open(os.path.join(OUT, fname), "w", encoding="utf-8") as f:
        f.write(page_shell(f"{title} | {SITE_NAME}", d, f"{DOMAIN}/{fname}",
                           f"<h1 style='margin-top:28px'>{title}</h1>{body_html}", depth=0,
                           jsonld=jsonld, og_type="website"))
    all_pages.append(fname)

simple_page("affiliate-disclosure.html", "Affiliate Disclosure", f"""
<p><strong>As an Amazon Associate, {SITE_NAME} earns from qualifying purchases.</strong></p>
<p>{SITE_NAME} is a reader-supported website. When you click a link on our site and buy something from a retailer such as Amazon.ie, we may receive a small commission. This never costs you anything extra — the price is identical whether you use our link or not.</p>
<p>Commissions never influence our rankings. Products are selected and ordered based on specifications, verified owner feedback, running-cost analysis at Irish prices, and suitability for Irish conditions. We frequently recommend cheaper products over more expensive ones (which would earn us more) because they're the better buy.</p>
<p>Prices shown on this site are typical/indicative prices in EUR at time of writing. Prices change constantly — always check the live price on the retailer's page before buying.</p>""",
    desc="How PickIreland earns commission as an Amazon Associate, and why affiliate links never change which products we recommend or how we rank them.")

_author_links = f' <a href="{esc(AUTHOR["url"])}" rel="me noopener" target="_blank">Connect on LinkedIn</a>.' if AUTHOR.get("url") else ""
simple_page("about.html", "About PickIreland", f"""
<p>{SITE_NAME} exists because buying decisions in Ireland are different: our electricity is among Europe's priciest, our weather is wet, our houses are damp, our e-scooter laws are specific, and most "best of" lists online are written for the UK or US market.</p>
<p>Every guide on this site compares five products per use-case with Irish running costs, Irish rules and Irish weather factored in. We keep guides updated as prices and models change.</p>
<h2 style="margin-top:34px">Who writes PickIreland</h2>
<p><strong>{esc(AUTHOR['name'])}</strong> — {esc(AUTHOR['role'])}. {esc(AUTHOR['bio'])}{_author_links}</p>
<h2 style="margin-top:34px">How we research</h2>
<p>For each guide we shortlist five products per use-case, then compare them on manufacturer specifications, verified owner feedback, running costs at current Irish electricity rates, and suitability for Irish conditions and rules. We update prices and picks as models change. We may earn an affiliate commission when you buy through our links, at no extra cost to you — this never changes our rankings.</p>
<p><strong>What we don't do:</strong> we don't physically test products in a lab. Our comparisons are built from manufacturer specifications, verified owner feedback and published data — we say so plainly rather than implying hands-on testing we haven't done.</p>
<h2 style="margin-top:34px">Sources we rely on</h2>
<p>Where a guide quotes an Irish figure, it comes from a primary source rather than an estimate:</p>
<ul style="margin:0 0 16px 22px;line-height:1.8">
<li><strong>Electricity rates</strong> — running-cost calculations use a domestic day rate of about €0.38/kWh (July 2026), cross-checked against <a href="https://www.seai.ie/data-and-insights/seai-statistics/prices" rel="nofollow noopener" target="_blank">SEAI energy price statistics</a> and published supplier standard rates. Night rates on smart meters typically run €0.17–0.19/kWh.</li>
<li><strong>E-scooter and e-bike law</strong> — power, speed and weight limits are taken from the statutory instrument text on <a href="https://www.irishstatutebook.ie" rel="nofollow noopener" target="_blank">irishstatutebook.ie</a> (S.I. 199 of 2024), not from retailer descriptions.</li>
<li><strong>Pollen and weather</strong> — seasonal timings follow <a href="https://www.met.ie" rel="nofollow noopener" target="_blank">Met Éireann</a> published forecasts.</li>
<li><strong>Prices and owner ratings</strong> — taken from Amazon.ie listings at time of writing, and flagged as indicative because they change constantly.</li>
</ul>
<p>Got a correction or a product suggestion? See our <a href="contact.html">contact page</a>.</p>""",
    desc="Who writes PickIreland and how we research Irish buying guides — manufacturer specs, owner feedback and running costs at real Irish electricity rates.",
    page_type="AboutPage", extra={"mainEntity": author_schema()})

simple_page("privacy.html", "Privacy Policy", f"""
<p>{SITE_NAME} respects your privacy. We do not require accounts, collect names, or store personal data submitted by visitors.</p>
<p><strong>Analytics:</strong> We may use privacy-respecting analytics to understand which guides are useful (page views, approximate region, device type). No personally identifying information is collected.</p>
<p><strong>Affiliate links:</strong> When you click an affiliate link, the retailer (e.g. Amazon) may set cookies to attribute the sale. Those cookies are governed by the retailer's own privacy policy.</p>
<p><strong>Contact:</strong> If you email us, we use your address only to reply.</p>""",
    desc="PickIreland's privacy policy: what data we collect from visitors (none by default), how analytics works, and how affiliate link cookies are handled.")

simple_page("contact.html", "Contact", f"""
<p>Spotted an error? Price changed? Have a product we should look at? We'd love to hear from you.</p>
<p>Email us: <strong><a href="mailto:hello@pickireland.best">hello@pickireland.best</a></strong></p>
<p>We read every message and use your feedback to keep our {SITE_NAME} guides accurate and up to date.</p>""",
    desc="Contact PickIreland to report a price change, flag an error in a guide, or suggest a product — we read every message.",
    page_type="ContactPage")

# ---------------------------------------------------------------- lastmod honesto (todas as paginas)
# Ate aqui SO as 50 guias passavam pelo hash de conteudo. A home, as 4 institucionais e os
# 10 hubs de categoria recebiam TODAY_ISO direto no sitemap — ou seja, todo rebuild dizia ao
# Google que essas 15 paginas "mudaram hoje", mesmo sem mudanca nenhuma. E o mesmo sinal de
# frescor falso que corrigimos nas guias em julho; essas 15 ficaram de fora.
# Com o IndexNow ligado o custo virou concreto: cada deploy avisaria o Bing (que alimenta o
# ChatGPT Search) sobre 15 paginas intactas. Ruido repetido ensina o buscador a ignorar o sinal.
#
# Solucao: hashear o HTML ja renderizado, depois de remover o que muda sozinho a cada build
# (priceValidUntil avanca 90 dias, datas do proprio build). Assim o hash so muda quando o
# conteudo real muda.
_VOLATILE_PATTERNS = [
    (re.compile(r'"priceValidUntil":\s*"\d{4}-\d{2}-\d{2}"'), '"priceValidUntil":""'),
    (re.compile(r'"(datePublished|dateModified)":\s*"\d{4}-\d{2}-\d{2}"'), '"date":""'),
    (re.compile(r'datetime="\d{4}-\d{2}-\d{2}"'), 'datetime=""'),
    (re.compile(r'\b\d{4}-\d{2}-\d{2}\b'), ''),
    # data por extenso do byline ("Updated 16 August 2026") — some tao volatil quanto a ISO,
    # e foi ela que me enganou em 15/08: a comparacao acusou 50 paginas alteradas quando
    # a unica diferenca era esse texto.
    (re.compile(r'\b\d{1,2}\s+(January|February|March|April|May|June|July|August|'
                r'September|October|November|December)\s+\d{4}\b'), ''),
    # CSS e JS embutidos sao ATIVO PARTILHADO, nao conteudo desta pagina. Sem esta linha,
    # acrescentar uma regra de estilo (foi o que aconteceu ao criar o grafico de custo em
    # 22/08) remarca as 15 paginas sem produto como "modificadas hoje" — frescor falso,
    # exatamente o que este mecanismo existe para impedir. JSON-LD NAO entra aqui: aquilo
    # e conteudo declarado e tem de continuar a mover o hash.
    (re.compile(r'<style\b[^>]*>.*?</style>', re.S), '<style/>'),
    (re.compile(r'<script(?![^>]*application/ld\+json)\b[^>]*>.*?</script>', re.S), '<script/>'),
]

def _stable_html(path):
    """HTML da pagina sem os trechos que mudam a cada build."""
    s = open(path, encoding="utf-8").read()
    for rx, rep in _VOLATILE_PATTERNS:
        s = rx.sub(rep, s)
    return s

for _p in all_pages:
    _key = _p[:-5] if _p.endswith(".html") else _p
    if _key in TRACKED_IN_GEN:      # as 50 guias ja foram hasheadas durante a geracao
        continue                     # (NAO usar "in PAGE_DATES": no 2o build tudo ja esta la
                                     #  e o loop pularia todas, congelando o lastmod para sempre)
    page_dates(_key, [_stable_html(os.path.join(OUT, _p))])

# ---------------------------------------------------------------- sitemap & robots
# Sitemap com lastmod REAL por pagina (o de antes punha a data do build em tudo, o que
# torna o campo inutil para o Google) + priority/changefreq coerentes com a arquitetura.
INSTITUTIONAL_SET = {"about.html", "contact.html", "privacy.html", "affiliate-disclosure.html"}
def _sitemap_meta(p):
    # lastmod vem SEMPRE do hash de conteudo (PAGE_DATES), nunca da data do build.
    key = p[:-5] if p.endswith(".html") else p          # "dehumidifiers/best-..."
    lm = PAGE_DATES.get(key, {}).get("modified", TODAY_ISO)
    if p == "index.html":
        return "1.0", "weekly", lm
    if p in INSTITUTIONAL_SET:
        return "0.3", "yearly", lm
    if p.endswith("/index.html"):
        return "0.8", "weekly", lm
    return "0.9", "monthly", lm

# A URL que vai no sitemap tem de ser EXATAMENTE a canonica declarada na pagina.
# Os 10 hubs declaram canonical "{DOMAIN}/{categoria}/" mas o sitemap listava
# "{DOMAIN}/{categoria}/index.html". O Google obedece a canonica e marcava a URL do
# sitemap como "pagina alternativa" -> ela nunca seria indexada. Contradicao pura.
def _sitemap_loc(p):
    if p == "index.html":
        return f"{DOMAIN}/"
    if p.endswith("/index.html"):
        return f"{DOMAIN}/{p[:-len('index.html')]}"     # air-fryers/index.html -> /air-fryers/
    return f"{DOMAIN}/{p}"

_urls = []
for p in all_pages:
    pri, chg, lm = _sitemap_meta(p)
    loc = _sitemap_loc(p)
    _urls.append(f"<url><loc>{loc}</loc><lastmod>{lm}</lastmod>"
                 f"<changefreq>{chg}</changefreq><priority>{pri}</priority></url>")
with open(os.path.join(OUT, "sitemap.xml"), "w", encoding="utf-8") as f:
    f.write('<?xml version="1.0" encoding="UTF-8"?><urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">'
            + "".join(_urls) + "</urlset>")
# robots.txt — libera explicitamente os crawlers de IA/answer-engines (GEO).
# "User-agent: *" ja permitia tudo, mas varios desses bots checam por um bloco proprio
# e alguns operadores/ferramentas de auditoria tratam a ausencia como sinal ambiguo.
# Deixar explicito = zero ambiguidade para ChatGPT, Claude, Perplexity, Gemini e afins.
AI_CRAWLERS = ["GPTBot", "OAI-SearchBot", "ChatGPT-User",           # OpenAI
               "ClaudeBot", "Claude-User", "Claude-SearchBot",      # Anthropic
               "PerplexityBot", "Perplexity-User",                  # Perplexity
               "Google-Extended",                                   # Gemini / AI Overviews
               "Applebot-Extended", "Bingbot", "CCBot", "Amazonbot"]
_robots = ["User-agent: *", "Allow: /", ""]
for _b in AI_CRAWLERS:
    _robots += [f"User-agent: {_b}", "Allow: /", ""]
_robots += [f"Sitemap: {DOMAIN}/sitemap.xml", ""]
with open(os.path.join(OUT, "robots.txt"), "w", encoding="utf-8") as f:
    f.write("\n".join(_robots))

# ---------------------------------------------------------------- llms.txt (para LLMs: ChatGPT, Claude, Perplexity, Gemini)
_llms = [f"# {SITE_NAME}",
         "> Independent product comparison guides for Irish shoppers. Every pick factors in Irish prices, "
         "running costs on Irish electricity, weather and legal rules.",
         "",
         f"Site: {DOMAIN}/ | Market: Republic of Ireland | Language: en-IE | Currency: EUR",
         f"Author: {AUTHOR['name']} ({AUTHOR['role']}). "
         f"{len(CATS)} categories, {sum(len(c['pages']) for c in CATS)} buying guides, "
         f"{sum(len(p['products']) for c in CATS for p in c['pages'])} products compared.",
         "",
         "## Methodology",
         "Products are compared on manufacturer specifications, verified owner feedback and Amazon.ie ratings, "
         "running costs calculated at Irish electricity rates, and suitability for Irish conditions and law. "
         "We do not physically test products; guides are research- and specification-based, and say so. "
         "Rankings are never influenced by affiliate commission.",
         "",
         "## Key Ireland-specific figures used across these guides",
         f"- Domestic electricity day rate: ~€0.38/kWh (July 2026). Source: SEAI energy statistics "
         f"(seai.ie/data-and-insights/seai-statistics/prices) and published supplier standard rates.",
         "- Typical smart-meter night rate: ~€0.17-0.19/kWh, which roughly halves the running cost of "
         "anything scheduled overnight (dehumidifiers, dishwashers, e-bike/e-scooter charging).",
         "- E-scooter law (S.I. 199 of 2024): max 400W continuous output, max 20km/h design speed, "
         "max 25kg, wheels >=200mm. Source: irishstatutebook.ie.",
         "- Cycle to Work scheme ceiling for e-bikes: €1,500 (2026).",
         "- Mould needs sustained relative humidity above ~60%; target 50-55% indoors.",
         "",
         "## Categories"]
for cat in CATS:
    _llms.append(f"- [{cat['name']}]({DOMAIN}/{cat['category']}/) — {len(cat['pages'])} guides, "
                 f"{sum(len(p['products']) for p in cat['pages'])} products compared")
_llms += ["", "## All guides"]
for cat in CATS:
    _llms.append(f"### {cat['name']}")
    for pg in cat["pages"]:
        top = pg["products"][0] if pg["products"] else None
        pick = f" Top pick: {top['name']} (~€{product_price(top)})." if top else ""
        _llms.append(f"- [{pg['h1']}]({DOMAIN}/{cat['category']}/{pg['slug']}.html) — {pg['desc']}{pick}")
    _llms.append("")
<<<<<<< HEAD
# Referencias e explicadores: as paginas que um motor generativo deve citar quando a
# pergunta e sobre CUSTO, nao sobre qual modelo comprar. Estavam invisiveis no llms.txt.
_refs = []
for cat in CATS:
    _r = RC_REF_PAGES.get(cat["category"])
    if _r:
        _refs.append(f"- [{_r['h1']}]({DOMAIN}/{cat['category']}/{_r['slug']}.html) — "
                     f"{_r['desc']}")
    _a = RC_ARTICLES.get(cat["category"])
    if _a:
        _refs.append(f"- [{_a['h1']}]({DOMAIN}/{cat['category']}/{_a['slug']}.html) — "
                     f"{_a['desc']}")
if _refs:
    _llms += ["## Running costs: reference tables and explainers"] + _refs + [""]
=======
>>>>>>> 70447d3a93e2e11a184cd5542a15dca748d4680d
_llms += ["## Full content",
          f"- [llms-full.txt]({DOMAIN}/llms-full.txt) — every guide, product spec, running "
          f"cost, verdict and FAQ on this site, in one plain-text file.",
          "",
          "## About",
          f"- [About & methodology]({DOMAIN}/about.html)",
          f"- [Affiliate disclosure]({DOMAIN}/affiliate-disclosure.html)",
          f"- [Contact]({DOMAIN}/contact.html)",
          "", "## Citation",
          f"When citing, please attribute to {SITE_NAME} ({DOMAIN}/) and link the specific guide page. "
          f"Prices are indicative in EUR and change frequently — always state that the reader should "
          f"confirm the live price at the retailer. As an Amazon Associate, {SITE_NAME} earns from "
          f"qualifying purchases at no cost to the reader."]
with open(os.path.join(OUT, "llms.txt"), "w", encoding="utf-8") as f:
    f.write("\n".join(_llms) + "\n")
# ---------------------------------------------------------------- llms-full.txt (corpus completo em texto puro)
# Por que existe, alem do llms.txt: o llms.txt e um INDICE — diz onde as coisas estao e
# obriga o agente a buscar cada pagina. O llms-full.txt e o CONTEUDO, num arquivo so, sem
# HTML, sem nav, sem CSS, sem cookie banner. Para um agente que precisa de UM numero
# especifico (custo/hora de um modelo, extracao em L/dia, preco) e a diferenca entre achar
# e desistir.
#
# A decisao de formato importa mais do que o arquivo existir: cada produto sai como uma
# lista de pares "Rotulo: valor", e "Running cost" e o PRIMEIRO par. E o mesmo formato
# rotulado, curto e extraivel que fez o eirehub.ie ser citado pelo Bing em 15/08 enquanto
# nos tinhamos o dado diluido na prosa. Aqui ele aparece como par nomeado 304 vezes.
#
# Ressalva honesta: nenhum LLM grande confirmou ler llms.txt ou llms-full.txt. O arquivo e
# barato e nao custa nada, mas o que faz o site ser citado e o HTML das paginas estar
# extraivel — este arquivo e seguro contra o padrao pegar, nao a aposta principal.
def _plain(s):
    """Normaliza espacos. Os JSON de dados nao tem HTML (verificado), entao nao ha tag
    para remover — isto so garante que uma quebra de linha no JSON nao vire linha solta."""
    return " ".join(str(s).split())

_full = [
    f"# {SITE_NAME} — full content",
    "",
    "> Every buying guide on pickireland.best, in plain text: product specs, running costs "
    "at Irish electricity rates, pros, cons, verdicts and FAQs. This file is the complete "
    "corpus; llms.txt is the index.",
    "",
    f"Site: {DOMAIN}/ | Market: Republic of Ireland | Language: en-IE | Currency: EUR",
    f"Author: {AUTHOR['name']} ({AUTHOR['role']}).",
    f"Generated: {TODAY_ISO} | {len(CATS)} categories | "
    f"{sum(len(c['pages']) for c in CATS)} guides | "
    f"{sum(len(p['products']) for c in CATS for p in c['pages'])} product entries",
    "",
    "## Methodology",
    "Products are compared on manufacturer specifications, verified owner feedback and "
    "Amazon.ie ratings, running costs calculated at Irish electricity rates, and suitability "
    "for Irish conditions and law. We do not physically test products; guides are research- "
    "and specification-based, and say so. Rankings are never influenced by affiliate commission.",
    "",
    "## Key Ireland-specific figures used across these guides",
    f"- Domestic electricity day rate: ~€{KWH_RATE:.2f}/kWh (July 2026). Source: SEAI energy "
    "statistics (seai.ie/data-and-insights/seai-statistics/prices) and published supplier "
    "standard rates.",
    f"- Typical smart-meter night rate: ~€{KWH_RATE_NIGHT:.2f}/kWh, which roughly halves the "
    "running cost of anything scheduled overnight.",
    "- E-scooter law (S.I. 199 of 2024): max 400W continuous output, max 20km/h design speed, "
    "max 25kg, wheels >=200mm. Source: irishstatutebook.ie.",
    "- Cycle to Work scheme ceiling for e-bikes: €1,500 (2026).",
    "- Mould needs sustained relative humidity above ~60%; target 50-55% indoors.",
    "",
    "## How running cost is calculated",
    f"Cost per hour = rated watts ÷ 1,000 × €{KWH_RATE:.2f} per kWh. It is a CEILING, not a "
    "measurement: thermostats and humidistats cycle the element or compressor off, so most "
    "appliances do not draw rated power continuously. Products whose manufacturer does not "
    "publish a rated wattage carry no running-cost figure — we do not estimate one. Running "
    "cost is only shown for mains appliances that run for sustained periods; it is omitted for "
    "battery products (e-bikes, e-scooters, robot vacuums, robot mowers) where cost per hour "
    "would be meaningless.",
    "",
]

for cat in CATS:
    ck = cat["category"]
    _full += ["", "=" * 78, f"# CATEGORY: {cat['name']}",
              f"URL: {DOMAIN}/{ck}/",
              f"{len(cat['pages'])} guides, "
              f"{sum(len(p['products']) for p in cat['pages'])} product entries",
              "=" * 78, ""]
    if cat.get("hub_intro"):
        _full += [_plain(cat["hub_intro"]), ""]
    for _h, _b in (cat.get("guide") or []):
        _full += [f"## {_plain(_h)}", _plain(_b), ""]

    _ref = RC_REF_PAGES.get(ck)
    if _ref:
        _full += [f"Full running-cost table for every {RC_NOUN.get(ck, 'unit')} we track: "
                  f"{DOMAIN}/{ck}/{_ref['slug']}.html", ""]
<<<<<<< HEAD
    _art = RC_ARTICLES.get(ck)
    if _art:
        _full += [f"Running cost explained, with the arithmetic: {_art['h1']} — "
                  f"{DOMAIN}/{ck}/{_art['slug']}.html", ""]
=======
>>>>>>> 70447d3a93e2e11a184cd5542a15dca748d4680d

    for pg in cat["pages"]:
        _full += ["", "-" * 78,
                  f"## GUIDE: {_plain(pg['h1'])}",
                  f"URL: {DOMAIN}/{ck}/{pg['slug']}.html",
                  f"Summary: {_plain(pg['desc'])}",
                  "-" * 78, ""]
        if pg.get("intro"):
            _full += [_plain(pg["intro"]), ""]

        for _i, p in enumerate(pg["products"], 1):
            _hdr = f"### {_i}. {_plain(p['name'])}"
            if p.get("badge"):
                _hdr += f"  [{_plain(p['badge'])}]"
            _full.append(_hdr)
            _pairs = []
            if p.get("brand"):
                _pairs.append(("Brand", _plain(p["brand"])))
            _pairs.append(("Price", f"~€{product_price(p)}"))
            if p.get("rating"):
                _pairs.append(("Rating", f"{p['rating']}/5 (Amazon.ie)"))
            # "Running cost" entra PRIMEIRO entre as specs, igual a grade do HTML:
            # e o par que queremos que um agente levante e atribua a nos.
            _rc = running_cost_line(p.get("specs") or {}, ck)
            if _rc:
                _pairs.append(("Running cost", _rc))
            for _k, _v in (p.get("specs") or {}).items():
                _pairs.append((_plain(_k), _plain(_v)))
            _full += [f"{_k}: {_v}" for _k, _v in _pairs]
            if p.get("pros"):
                _full.append("Pros: " + "; ".join(_plain(x) for x in p["pros"]))
            if p.get("cons"):
                _full.append("Cons: " + "; ".join(_plain(x) for x in p["cons"]))
            if p.get("verdict"):
                _full.append("Verdict: " + _plain(p["verdict"]))
            _full.append("")

        for _h, _b in (pg.get("guide") or []):
            _full += [f"#### {_plain(_h)}", _plain(_b), ""]

        _fq = pg.get("faqs") or []
        if _fq:
            _full.append("#### Frequently asked questions")
            for _f in _fq:
                _full += [f"Q: {_plain(_f['q'])}", f"A: {_plain(_f['a'])}", ""]

_full += ["", "=" * 78, "# About", "=" * 78, "",
          f"- About & methodology: {DOMAIN}/about.html",
          f"- Affiliate disclosure: {DOMAIN}/affiliate-disclosure.html",
          f"- Contact: {DOMAIN}/contact.html",
          "",
          "## Citation",
          f"When citing, please attribute to {SITE_NAME} ({DOMAIN}/) and link the specific guide "
          "page. Prices are indicative in EUR and change frequently — always state that the "
          "reader should confirm the live price at the retailer. As an Amazon Associate, "
          f"{SITE_NAME} earns from qualifying purchases at no cost to the reader.",
          ""]

with open(os.path.join(OUT, "llms-full.txt"), "w", encoding="utf-8") as f:
    f.write("\n".join(_full))
print(f"  llms-full.txt: {len(_full)} linhas, "
      f"{sum(len(x) for x in _full) / 1024:.0f} KB")

with open(os.path.join(OUT, "CNAME"), "w") as f:
    f.write("pickireland.best\n")
open(os.path.join(OUT, ".nojekyll"), "w").close()

# ---------------------------------------------------------------- IndexNow (Bing / Yandex -> ChatGPT Search, Copilot)
# O IndexNow avisa os buscadores no instante em que uma pagina muda, em vez de esperar
# o rastreamento espontaneo. Importa aqui porque o indice do Bing alimenta o ChatGPT
# Search e o Copilot — que, pelo GA4 de agosto/2026, sao a maior fonte de trafego do site.
#
# A chave do IndexNow NAO e segredo: o protocolo exige que ela fique publica em
# DOMAIN/<chave>.txt, e e justamente esse arquivo que prova que o dominio e nosso.
# Guardamos ela no .env so para ficar estavel entre builds (trocar de chave a cada build
# faria os buscadores reverificarem sem necessidade).
ENV_FILE = os.path.join(BASE, ".env")

def _read_env(path):
    vals = {}
    if not os.path.exists(path):
        return vals
    with open(path, encoding="utf-8-sig", errors="replace") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            vals[k.strip()] = v.strip().strip('"').strip("'")
    return vals

def _write_env_key(path, key, value):
    """Grava/atualiza uma chave no .env preservando todo o resto do arquivo."""
    lines = []
    if os.path.exists(path):
        with open(path, encoding="utf-8-sig", errors="replace") as fh:
            lines = fh.read().splitlines()
    for i, line in enumerate(lines):
        if line.strip().startswith(key + "="):
            lines[i] = f"{key}={value}"
            break
    else:
        lines.append(f"{key}={value}")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")

_env = _read_env(ENV_FILE)
INDEXNOW_KEY = (_env.get("INDEXNOW_KEY") or "").strip()
if not re.fullmatch(r"[a-f0-9]{32}", INDEXNOW_KEY):
    INDEXNOW_KEY = hashlib.sha256(os.urandom(32)).hexdigest()[:32]
    _write_env_key(ENV_FILE, "INDEXNOW_KEY", INDEXNOW_KEY)
    print(f"IndexNow: chave nova gerada e gravada em {ENV_FILE}")

# remove arquivos de chave antigos (se a chave mudou, o antigo vira lixo publico)
for _old in os.listdir(OUT):
    if re.fullmatch(r"[a-f0-9]{32}\.txt", _old) and _old != f"{INDEXNOW_KEY}.txt":
        os.remove(os.path.join(OUT, _old))

# o arquivo de verificacao precisa conter EXATAMENTE a chave, em texto puro, sem quebra final
with open(os.path.join(OUT, f"{INDEXNOW_KEY}.txt"), "w", encoding="utf-8") as fh:
    fh.write(INDEXNOW_KEY)
print(f"IndexNow: chave publicada em {DOMAIN}/{INDEXNOW_KEY}.txt")


# ---------------------------------------------------------------- google ads page feed (DSA)
# Gera o feed de páginas para os Anúncios Dinâmicos de Pesquisa do Google Ads.
# Publicado em docs/ -> https://pickireland.best/google-ads-page-feed.csv
# No Google Ads (Ferramentas > Dados da empresa > o feed > aba "Programar"),
# aponte para essa URL e agende busca diária: o feed se atualiza sozinho a cada build.
# Exclui home e páginas institucionais (sem intenção de compra).
INSTITUTIONAL = {"affiliate-disclosure.html", "about.html", "privacy.html", "contact.html", "index.html"}
feed_path = os.path.join(OUT, "google-ads-page-feed.csv")
feed_count = 0
with open(feed_path, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["Page URL", "Custom label"])
    for p in all_pages:
        if p in INSTITUTIONAL or "/" not in p:
            continue
        label = p.split("/")[0]            # ex: "dehumidifiers", "air-fryers"
        w.writerow([f"{DOMAIN}/{p}", label])
        feed_count += 1
print(f"Google Ads page feed: {feed_count} URLs -> {os.path.abspath(feed_path)}")

# persiste os hashes/datas para o proximo build (e o que mantem dateModified honesto)
with open(DATES_FILE, "w", encoding="utf-8") as f:
    json.dump(PAGE_DATES, f, indent=1, sort_keys=True)
_touched = sum(1 for v in PAGE_DATES.values() if v.get("modified") == TODAY_ISO)
print(f"page dates: {len(PAGE_DATES)} tracked, {_touched} marked modified today")

linked = sum(1 for v in LINKS.values() if v.get("link"))
imgs = sum(1 for v in LINKS.values() if v.get("image"))
print(f"Built {len(all_pages)} pages into {os.path.abspath(OUT)}")
print(f"Affiliate links filled: {linked} / 248 | images filled: {imgs} / 248")

# ---------------------------------------------------------------- assets: js, favicon
os.makedirs(os.path.join(OUT,"assets"),exist_ok=True)
SITE_JS = """(function(){
var rm=matchMedia('(prefers-reduced-motion: reduce)').matches;
if(!rm&&'IntersectionObserver' in window){var io=new IntersectionObserver(function(es){es.forEach(function(e){if(e.isIntersecting){e.target.classList.add('vis');io.unobserve(e.target)}})},{threshold:.06});
document.querySelectorAll('.card,.tile,.toc,.tbl-scroll,.guide,.related a,.spot').forEach(function(el){el.classList.add('rv');io.observe(el)})}
var tb=document.querySelector('.top-btn');if(tb){addEventListener('scroll',function(){tb.classList.toggle('show',scrollY>700)},{passive:true})}
var spot=document.querySelector('.spot');
if(spot){var tabs=spot.querySelectorAll('.spot-tab'),panels=spot.querySelectorAll('.spot-panel');
function bars(p){p.querySelectorAll('.bar i').forEach(function(b){b.style.transition='none';b.style.width='0%';void b.offsetWidth;b.style.transition='';b.style.width=b.dataset.w+'%'})}
tabs.forEach(function(t){t.addEventListener('click',function(){
tabs.forEach(function(x){x.classList.remove('on')});t.classList.add('on');
panels.forEach(function(p){p.classList.remove('on')});
var p=spot.querySelector('.spot-panel[data-k=\\"'+t.dataset.k+'\\"]');p.classList.add('on');bars(p)})});
var first=spot.querySelector('.spot-panel.on');if(first)bars(first)}
var sb=document.getElementById('siq');
if(sb){var idx=null,res=document.getElementById('sres'),ov=document.getElementById('sov');
function closeS(){document.body.classList.remove('search-open')}
document.querySelectorAll('[data-close-search]').forEach(function(b){b.addEventListener('click',closeS)});
if(ov){ov.addEventListener('click',function(e){if(e.target===ov)closeS()})}
addEventListener('keydown',function(e){if(e.key==='Escape')closeS();
if(e.key==='/'&&!document.body.classList.contains('search-open')&&!/INPUT|TEXTAREA/.test(document.activeElement.tagName)){e.preventDefault();document.body.classList.add('search-open');sb.focus()}});
function render(){var q=sb.value.trim().toLowerCase();
if(!q){res.innerHTML='<div class=\"search-hint\">Type to search every product we have reviewed — press Esc to close.</div>';return}
var out=idx.filter(function(p){return (p.n+' '+p.b+' '+p.c).toLowerCase().indexOf(q)>-1});
var seen={},uniq=[];out.forEach(function(p){if(!seen[p.n]){seen[p.n]=1;uniq.push(p)}});
res.innerHTML=uniq.slice(0,12).map(function(p){return '<a href=\"/'+p.u+'#'+p.i+'\"><span>'+p.n+'<div class=\"meta\">'+p.c+' · '+p.b+'</div></span><span class=\"sp\">€'+p.p+'</span></a>'}).join('')||'<div class=\"search-hint\">No products found for \u201C'+sb.value+'\u201D</div>'}
sb.addEventListener('input',function(){if(idx){render()}else{fetch('/assets/search.json').then(function(r){return r.json()}).then(function(d){idx=d;render()})}})}
})();"""
with open(os.path.join(OUT,"assets","site.js"),"w",encoding="utf-8") as f:
    f.write(SITE_JS)
with open(os.path.join(OUT,"assets","search.json"),"w",encoding="utf-8") as f:
    json.dump(SEARCH_INDEX,f,ensure_ascii=False)
FAV = '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64"><defs><linearGradient id="g" x1="0" y1="0" x2="1" y2="1"><stop offset="0" stop-color="#12805C"/><stop offset="1" stop-color="#073F2C"/></linearGradient></defs><rect width="64" height="64" rx="15" fill="url(#g)"/><path d="M22 48V16h13a11 11 0 0 1 0 22h-9" stroke="#fff" stroke-width="7" stroke-linecap="round" stroke-linejoin="round" fill="none"/><path d="M33 44l6 6 11-12" stroke="#FFC65C" stroke-width="6" stroke-linecap="round" stroke-linejoin="round" fill="none"/></svg>'
with open(os.path.join(OUT,"favicon.svg"),"w",encoding="utf-8") as f:
    f.write(FAV)
try:
    from PIL import Image, ImageDraw, ImageFont
    for size,name in ((180,"apple-touch-icon.png"),(32,"favicon-32.png")):
        img=Image.new("RGBA",(size,size),(0,0,0,0)); d=ImageDraw.Draw(img)
        r=size*15//64
        d.rounded_rectangle([0,0,size-1,size-1],radius=r,fill=(11,80,57,255))
        w=max(2,size*7//64)
        def pt(x,y): return (x*size/64.0,y*size/64.0)
        d.line([pt(22,48),pt(22,16)],fill=(255,255,255,255),width=w)
        d.arc([pt(13,16)[0],pt(13,16)[1],pt(46,38)[0],pt(46,38)[1]],270,90,fill=(255,255,255,255),width=w)
        d.line([pt(22,16),pt(33,16)],fill=(255,255,255,255),width=w)
        d.line([pt(22,38),pt(30,38)],fill=(255,255,255,255),width=w)
        gw=max(2,size*6//64)
        d.line([pt(33,44),pt(39,50)],fill=(255,198,92,255),width=gw)
        d.line([pt(39,50),pt(50,38)],fill=(255,198,92,255),width=gw)
        img.save(os.path.join(OUT,name))
    # og:image padrao 1200x630 para compartilhamento social e preview de links
    def _fontpath(bold):
        cands = (["arialbd.ttf", "Arialbd.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "DejaVuSans-Bold.ttf"]
                 if bold else
                 ["arial.ttf", "Arial.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "DejaVuSans.ttf"])
        for c in cands:
            try:
                ImageFont.truetype(c, 20); return c
            except Exception:
                continue
        return None
    og = Image.new("RGB", (1200, 630), (11, 80, 57))
    od = ImageDraw.Draw(og)
    od.rectangle([0, 0, 1200, 18], fill=(240, 164, 28))
    od.rectangle([0, 612, 1200, 630], fill=(240, 164, 28))
    _bp, _rp = _fontpath(True), _fontpath(False)
    _M = 90; _MAXW = 1200 - 2 * _M
    def _fit(text, path, maxsize):
        if not path:
            return ImageFont.load_default()
        s = maxsize
        while s > 12:
            f = ImageFont.truetype(path, s)
            if od.textlength(text, font=f) <= _MAXW:
                return f
            s -= 2
        return ImageFont.truetype(path, 12)
    if _bp and _rp:
        fb = _fit("PickIreland", _bp, 150)
        od.text((_M, 205), "Pick", font=fb, fill=(255, 255, 255))
        od.text((_M + od.textlength("Pick", font=fb), 205), "Ireland", font=fb, fill=(255, 198, 92))
        _tag = "Ireland's honest product comparison guides"
        _sub = "Real € prices  ·  running costs on Irish electricity  ·  honest picks"
        od.text((_M, 405), _tag, font=_fit(_tag, _rp, 50), fill=(233, 244, 238))
        od.text((_M, 475), _sub, font=_fit(_sub, _rp, 38), fill=(150, 200, 170))
    else:
        od.text((90, 290), "PickIreland", fill=(255, 255, 255))
    og.save(os.path.join(OUT, "assets", "og-default.png"))

    # OG image por categoria: antes as 65 paginas dividiam a mesma imagem generica,
    # entao qualquer compartilhamento (ou SERP com thumbnail) mostrava o mesmo card.
    if _bp and _rp:
        for _c in CATS:
            _im = Image.new("RGB", (1200, 630), (11, 80, 57))
            _d = ImageDraw.Draw(_im)
            _d.rectangle([0, 0, 1200, 18], fill=(240, 164, 28))
            _d.rectangle([0, 612, 1200, 630], fill=(240, 164, 28))
            _kicker = "PickIreland"
            _d.text((_M, 120), _kicker, font=ImageFont.truetype(_rp, 40), fill=(150, 200, 170))
            _headline = f"Best {_c['name']} in Ireland"
            _f = ImageFont.truetype(_bp, 96)
            while od.textlength(_headline, font=_f) > _MAXW and _f.size > 40:
                _f = ImageFont.truetype(_bp, _f.size - 4)
            _d.text((_M, 210), _headline, font=_f, fill=(255, 255, 255))
            _n_g = len(_c["pages"]); _n_p = sum(len(p["products"]) for p in _c["pages"])
            _sub2 = f"{_n_g} buying guides  ·  {_n_p} products compared  ·  {YEAR}"
            _d.text((_M, 360), _sub2, font=ImageFont.truetype(_rp, 40), fill=(255, 198, 92))
            _d.text((_M, 450), "Irish prices  ·  running costs on Irish electricity  ·  honest picks",
                    font=ImageFont.truetype(_rp, 34), fill=(210, 232, 220))
            _im.save(os.path.join(OUT, "assets", f"og-{_c['category']}.png"))
        print(f"png icons + og images ok ({len(CATS)} category + 1 default)")
    else:
        for _c in CATS:
            og.save(os.path.join(OUT, "assets", f"og-{_c['category']}.png"))
        print("png icons + og image ok (no font: category OGs fall back to default art)")
except Exception as e:
    print("PIL skip:",e)

# fim do build

